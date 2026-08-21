#!/usr/bin/env python3
"""
Generate Excel simulation file for MW Planner Recommendation Engine V2
Creates 200 inventories with formulas for scenario testing
"""

import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule

# Constants
COUNTRIES = ['Malaysia', 'Singapore', 'Indonesia', 'Thailand', 'Philippines']
CITIES = {
    'Malaysia': ['Kuala Lumpur', 'Penang', 'Johor Bahru', 'KLIA/Sepang', 'Melaka', 'Ipoh', 'Kuching', 'Kota Kinabalu'],
    'Singapore': ['Singapore Central', 'Orchard', 'Marina Bay', 'Changi', 'Jurong'],
    'Indonesia': ['Jakarta', 'Surabaya', 'Bandung', 'Bali', 'Medan'],
    'Thailand': ['Bangkok', 'Phuket', 'Chiang Mai', 'Pattaya'],
    'Philippines': ['Manila', 'Cebu', 'Davao', 'Clark']
}
CITY_TIERS = {
    'Kuala Lumpur': 1, 'Singapore Central': 1, 'Orchard': 1, 'Jakarta': 1, 'Bangkok': 1, 'Manila': 1,
    'Penang': 2, 'Johor Bahru': 2, 'KLIA/Sepang': 2, 'Marina Bay': 2, 'Changi': 2, 'Surabaya': 2, 'Bandung': 2,
    'Bali': 2, 'Phuket': 2, 'Cebu': 2,
}
FORMATS = ['Digital Screen', 'Classic', 'Transit', 'Retail', 'Network', 'Radio', 'Experiential']
FORMAT_WEIGHTS = [0.25, 0.20, 0.20, 0.15, 0.12, 0.04, 0.04]
VENUES = {
    'Digital Screen': ['LED Billboard', 'Digital Tower', 'Mega Screen', 'Digital Wall'],
    'Classic': ['Static Billboard', 'Poster Site', 'Unipole', 'Hoarding'],
    'Transit': ['Bus Shelter', 'LRT Station', 'MRT Screen', 'Bus Wrap', 'Train Station'],
    'Retail': ['Mall Screen', 'Supermarket Display', 'Convenience Store', 'Food Court Screen'],
    'Network': ['Airport Network', 'Office Building', 'Hotel Lobby', 'Hospital Network'],
    'Radio': ['FM Radio Spot', 'Digital Radio'],
    'Experiential': ['Pop-up Activation', 'Interactive Kiosk', 'Ambient Installation']
}
GOAL_TYPES = ['Impressions', 'Reach', 'SOV', 'Ad Plays', 'Carbon Emission']

def generate_inventory_name(format_type, city, idx):
    venue = random.choice(VENUES[format_type])
    locations = {
        'Kuala Lumpur': ['KLCC', 'Pavilion', 'Bukit Bintang', 'KL Sentral', 'Bangsar', 'Mid Valley', 'Mont Kiara'],
        'Penang': ['Gurney', 'Komtar', 'Queensbay', 'Penang Airport', 'Georgetown'],
        'Johor Bahru': ['JB Sentral', 'City Square', 'Danga Bay', 'Senai', 'Larkin'],
        'KLIA/Sepang': ['KLIA T1', 'KLIA T2', 'Sepang F1', 'Gateway Mall'],
        'Melaka': ['Jonker Street', 'Dataran Pahlawan', 'Mahkota Parade', 'The Shore'],
        'Singapore Central': ['Orchard MRT', 'Raffles Place', 'City Hall', 'Bugis'],
        'Jakarta': ['Sudirman', 'SCBD', 'Thamrin', 'Menteng', 'Kemang'],
        'Bangkok': ['Siam', 'Sukhumvit', 'Silom', 'Chatuchak', 'Asok'],
        'Manila': ['Makati', 'BGC', 'Ortigas', 'MOA', 'Quezon City']
    }
    location = random.choice(locations.get(city, [city]))
    return f"{location} {venue}"

def get_base_metrics(format_type, city_tier):
    base = {
        'Digital Screen': {'imps': 45000, 'reach': 35000, 'cost': 35000, 'co2': 0.008},
        'Classic': {'imps': 25000, 'reach': 20000, 'cost': 8000, 'co2': 0.001},
        'Transit': {'imps': 15000, 'reach': 12000, 'cost': 5000, 'co2': 0.003},
        'Retail': {'imps': 8000, 'reach': 6000, 'cost': 4000, 'co2': 0.005},
        'Network': {'imps': 60000, 'reach': 45000, 'cost': 50000, 'co2': 0.006},
        'Radio': {'imps': 100000, 'reach': 80000, 'cost': 15000, 'co2': 0.0001},
        'Experiential': {'imps': 5000, 'reach': 4000, 'cost': 25000, 'co2': 0.010}
    }
    tier_multiplier = {1: 1.5, 2: 1.0, 3: 0.7}.get(city_tier, 0.8)
    metrics = base[format_type].copy()
    for key in ['imps', 'reach', 'cost']:
        metrics[key] = int(metrics[key] * tier_multiplier * random.uniform(0.7, 1.3))
    metrics['co2'] = metrics['co2'] * random.uniform(0.5, 1.5)
    return metrics

def create_workbook():
    wb = Workbook()
    
    # Sheet 1: Campaign Inputs
    ws_inputs = wb.active
    ws_inputs.title = "Campaign Inputs"
    
    # Style definitions
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Campaign Input Section
    inputs = [
        ("Campaign Inputs", None, "header"),
        ("Country", "Malaysia", "input"),
        ("Target City (leave blank for all)", "", "input"),
        ("Budget (local currency)", 100000, "input"),
        ("Campaign Start Date", "2026-01-01", "input"),
        ("Campaign End Date", "2026-01-31", "input"),
        ("Campaign Days", "=B6-B5+1", "formula"),
        ("Goal Type", "Impressions", "input"),
        ("Goal Value", 500000, "input"),
        ("Venue Type Preference (leave blank for all)", "", "input"),
        ("", None, None),
        ("Scoring Weights", None, "header"),
        ("measure_fit weight", 0.20, "input"),
        ("geo_fit weight", 0.20, "input"),
        ("availability weight", 0.10, "input"),
        ("budget_fit weight", 0.20, "input"),
        ("audience_fit weight", 0.10, "input"),
        ("brand_fit weight", 0.10, "input"),
        ("quality_fit weight", 0.06, "input"),
        ("time_fit weight", 0.04, "input"),
        ("", None, None),
        ("City Diversity Settings", None, "header"),
        ("Max single city %", 0.50, "input"),
        ("Min Tier 1 city %", 0.15, "input"),
        ("Min Tier 2 city %", 0.10, "input"),
        ("Min Tier 3 city %", 0.05, "input"),
        ("", None, None),
        ("Budget Tolerance", None, "header"),
        ("Tolerance %", 0.07, "input"),
    ]
    
    for i, (label, value, cell_type) in enumerate(inputs, 1):
        ws_inputs.cell(row=i, column=1, value=label)
        if value is not None:
            ws_inputs.cell(row=i, column=2, value=value)
        if cell_type == "header":
            ws_inputs.cell(row=i, column=1).font = Font(bold=True, size=12)
        elif cell_type == "input":
            ws_inputs.cell(row=i, column=2).fill = input_fill
            ws_inputs.cell(row=i, column=2).border = thin_border
    
    ws_inputs.column_dimensions['A'].width = 35
    ws_inputs.column_dimensions['B'].width = 20
    
    # Sheet 2: Inventory Data (200 items)
    ws_inv = wb.create_sheet("Inventory Data")
    
    inv_headers = [
        "ID", "Name", "Country", "City", "City Tier", "Format", "Venue Type",
        "Daily Impressions", "Daily Reach", "Daily Cost", "CO2/Play (kg)",
        "Operating Hours", "Min Days", "Availability %", "Quality Score",
        "Audience Match %", "Brand Fit %", "Loop Length (sec)", "Spot Length (sec)",
        "Plays/Hour"
    ]
    
    for col, header in enumerate(inv_headers, 1):
        cell = ws_inv.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    # Generate 200 inventories
    inventory_data = []
    inv_id = 1
    
    for country in COUNTRIES:
        cities = CITIES[country]
        inv_per_country = 40 if country == 'Malaysia' else 40
        
        for _ in range(inv_per_country):
            city = random.choice(cities)
            city_tier = CITY_TIERS.get(city, 3)
            format_type = random.choices(FORMATS, weights=FORMAT_WEIGHTS)[0]
            metrics = get_base_metrics(format_type, city_tier)
            venue_type = random.choice(VENUES[format_type])
            
            row_data = [
                inv_id,
                generate_inventory_name(format_type, city, inv_id),
                country,
                city,
                city_tier,
                format_type,
                venue_type,
                metrics['imps'],
                metrics['reach'],
                metrics['cost'],
                round(metrics['co2'], 5),
                random.choice([12, 14, 16, 18, 20, 24]),
                random.choice([1, 3, 5, 7, 14]),
                random.randint(70, 100),
                random.randint(50, 100),
                random.randint(40, 95),
                random.randint(40, 90),
                random.choice([30, 45, 60, 90, 120]),
                random.choice([10, 15, 20, 30]),
                60
            ]
            inventory_data.append(row_data)
            inv_id += 1
    
    for row_idx, row_data in enumerate(inventory_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws_inv.cell(row=row_idx, column=col_idx, value=value)
    
    for col in range(1, len(inv_headers) + 1):
        ws_inv.column_dimensions[get_column_letter(col)].width = 15
    ws_inv.column_dimensions['B'].width = 30
    
    # Sheet 3: Scoring Calculation
    ws_score = wb.create_sheet("Scoring")
    
    score_headers = [
        "ID", "Name", "City", "Format", "Country Match", "City Match", "Venue Match",
        "measure_fit", "geo_fit", "availability", "budget_fit",
        "audience_fit", "brand_fit", "quality_fit", "time_fit",
        "Total Score", "Rank"
    ]
    
    for col, header in enumerate(score_headers, 1):
        cell = ws_score.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    for row_idx in range(2, 202):
        inv_row = row_idx
        ws_score.cell(row=row_idx, column=1, value=f"='Inventory Data'!A{inv_row}")
        ws_score.cell(row=row_idx, column=2, value=f"='Inventory Data'!B{inv_row}")
        ws_score.cell(row=row_idx, column=3, value=f"='Inventory Data'!D{inv_row}")
        ws_score.cell(row=row_idx, column=4, value=f"='Inventory Data'!F{inv_row}")
        
        # Country match
        ws_score.cell(row=row_idx, column=5, value=f"=IF('Inventory Data'!C{inv_row}='Campaign Inputs'!$B$2,1,0)")
        
        # City match
        ws_score.cell(row=row_idx, column=6, value=f"=IF(OR('Campaign Inputs'!$B$3=\"\",'Inventory Data'!D{inv_row}='Campaign Inputs'!$B$3),1,0)")
        
        # Venue match (new column for venue preference filtering)
        venue_formula = f"""=IF(OR('Campaign Inputs'!$B$10="",ISNUMBER(SEARCH('Campaign Inputs'!$B$10,'Inventory Data'!G{inv_row}))),1,0)"""
        ws_score.cell(row=row_idx, column=7, value=venue_formula)
        
        # measure_fit - column 8
        measure_formula = f"""=IF(AND(E{row_idx}=1,G{row_idx}=1),
IF('Campaign Inputs'!$B$8="Impressions",MIN(100,'Inventory Data'!H{inv_row}*'Campaign Inputs'!$B$7/'Campaign Inputs'!$B$9*100),
IF('Campaign Inputs'!$B$8="Reach",MIN(100,'Inventory Data'!I{inv_row}*'Campaign Inputs'!$B$7/'Campaign Inputs'!$B$9*100),
IF('Campaign Inputs'!$B$8="SOV",80,
IF('Campaign Inputs'!$B$8="Ad Plays",MIN(100,'Inventory Data'!T{inv_row}*'Inventory Data'!L{inv_row}*'Campaign Inputs'!$B$7/'Campaign Inputs'!$B$9*100),
IF('Campaign Inputs'!$B$8="Carbon Emission",100-MIN(100,'Inventory Data'!K{inv_row}*1000),80))))),0)"""
        ws_score.cell(row=row_idx, column=8, value=measure_formula)
        
        # geo_fit - column 9
        ws_score.cell(row=row_idx, column=9, value=f"=IF(AND(F{row_idx}=1,G{row_idx}=1),IF('Inventory Data'!E{inv_row}=1,100,IF('Inventory Data'!E{inv_row}=2,80,60)),0)")
        
        # availability - column 10
        ws_score.cell(row=row_idx, column=10, value=f"=IF(G{row_idx}=1,'Inventory Data'!N{inv_row},0)")
        
        # budget_fit - column 11
        ws_score.cell(row=row_idx, column=11, value=f"=IF(G{row_idx}=1,IF('Inventory Data'!J{inv_row}*'Campaign Inputs'!$B$7<='Campaign Inputs'!$B$4,100,MAX(0,100-('Inventory Data'!J{inv_row}*'Campaign Inputs'!$B$7-'Campaign Inputs'!$B$4)/'Campaign Inputs'!$B$4*100)),0)")
        
        # audience_fit - column 12
        ws_score.cell(row=row_idx, column=12, value=f"=IF(G{row_idx}=1,'Inventory Data'!P{inv_row},0)")
        
        # brand_fit - column 13
        ws_score.cell(row=row_idx, column=13, value=f"=IF(G{row_idx}=1,'Inventory Data'!Q{inv_row},0)")
        
        # quality_fit - column 14
        ws_score.cell(row=row_idx, column=14, value=f"=IF(G{row_idx}=1,'Inventory Data'!O{inv_row},0)")
        
        # time_fit - column 15
        ws_score.cell(row=row_idx, column=15, value=f"=IF(G{row_idx}=1,MIN(100,'Inventory Data'!L{inv_row}/24*100),0)")
        
        # Total Score - column 16 (includes venue match filter)
        score_formula = f"""=IF(AND(E{row_idx}=1,G{row_idx}=1),
H{row_idx}*'Campaign Inputs'!$B$13+
I{row_idx}*'Campaign Inputs'!$B$14+
J{row_idx}*'Campaign Inputs'!$B$15+
K{row_idx}*'Campaign Inputs'!$B$16+
L{row_idx}*'Campaign Inputs'!$B$17+
M{row_idx}*'Campaign Inputs'!$B$18+
N{row_idx}*'Campaign Inputs'!$B$19+
O{row_idx}*'Campaign Inputs'!$B$20,0)"""
        ws_score.cell(row=row_idx, column=16, value=score_formula)
        
        # Rank - column 17
        ws_score.cell(row=row_idx, column=17, value=f"=IF(P{row_idx}=0,999,RANK(P{row_idx},$P$2:$P$201,0))")
    
    for col in range(1, len(score_headers) + 1):
        ws_score.column_dimensions[get_column_letter(col)].width = 14
    ws_score.column_dimensions['B'].width = 25
    
    # Sheet 4: Selection Results
    ws_results = wb.create_sheet("Selection Results")
    
    result_headers = [
        "Rank", "ID", "Name", "City", "Format", "Score", "Daily Cost", 
        "Daily Impressions", "Campaign Cost", "Campaign Impressions", 
        "Cumulative Budget", "Cumulative Impressions", "Within Budget?"
    ]
    
    for col, header in enumerate(result_headers, 1):
        cell = ws_results.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Add formulas to show top 30 inventories by rank
    for row_idx in range(2, 32):  # Top 30
        rank = row_idx - 1
        # Rank
        ws_results.cell(row=row_idx, column=1, value=rank)
        
        # ID - find the inventory with this rank (Rank is now in column Q)
        ws_results.cell(row=row_idx, column=2, 
            value=f'=IFERROR(INDEX(Scoring!A$2:A$201,MATCH({rank},Scoring!Q$2:Q$201,0)),"")')
        
        # Name
        ws_results.cell(row=row_idx, column=3,
            value=f'=IFERROR(INDEX(Scoring!B$2:B$201,MATCH({rank},Scoring!Q$2:Q$201,0)),"")')
        
        # City
        ws_results.cell(row=row_idx, column=4,
            value=f'=IFERROR(INDEX(Scoring!C$2:C$201,MATCH({rank},Scoring!Q$2:Q$201,0)),"")')
        
        # Format
        ws_results.cell(row=row_idx, column=5,
            value=f'=IFERROR(INDEX(Scoring!D$2:D$201,MATCH({rank},Scoring!Q$2:Q$201,0)),"")')
        
        # Score (Total Score is now in column P)
        ws_results.cell(row=row_idx, column=6,
            value=f'=IFERROR(INDEX(Scoring!P$2:P$201,MATCH({rank},Scoring!Q$2:Q$201,0)),"")')
        
        # Daily Cost - lookup from Inventory Data
        ws_results.cell(row=row_idx, column=7,
            value=f'=IFERROR(INDEX(\'Inventory Data\'!J$2:J$201,MATCH(B{row_idx},\'Inventory Data\'!A$2:A$201,0)),"")')
        
        # Daily Impressions
        ws_results.cell(row=row_idx, column=8,
            value=f'=IFERROR(INDEX(\'Inventory Data\'!H$2:H$201,MATCH(B{row_idx},\'Inventory Data\'!A$2:A$201,0)),"")')
        
        # Campaign Cost = Daily Cost * Campaign Days
        ws_results.cell(row=row_idx, column=9,
            value=f'=IF(B{row_idx}<>"",G{row_idx}*\'Campaign Inputs\'!$B$7,"")')
        
        # Campaign Impressions = Daily Impressions * Campaign Days
        ws_results.cell(row=row_idx, column=10,
            value=f'=IF(B{row_idx}<>"",H{row_idx}*\'Campaign Inputs\'!$B$7,"")')
        
        # Cumulative Budget
        if row_idx == 2:
            ws_results.cell(row=row_idx, column=11, value=f'=I{row_idx}')
        else:
            ws_results.cell(row=row_idx, column=11, value=f'=K{row_idx-1}+I{row_idx}')
        
        # Cumulative Impressions
        if row_idx == 2:
            ws_results.cell(row=row_idx, column=12, value=f'=J{row_idx}')
        else:
            ws_results.cell(row=row_idx, column=12, value=f'=L{row_idx-1}+J{row_idx}')
        
        # Within Budget?
        ws_results.cell(row=row_idx, column=13,
            value=f'=IF(K{row_idx}<=\'Campaign Inputs\'!$B$4,"YES","OVER")')
    
    # Add summary at bottom
    ws_results.cell(row=34, column=1, value="SUMMARY").font = Font(bold=True)
    ws_results.cell(row=35, column=1, value="Budget:")
    ws_results.cell(row=35, column=2, value="='Campaign Inputs'!B4")
    ws_results.cell(row=36, column=1, value="Selected within budget:")
    ws_results.cell(row=36, column=2, value='=COUNTIF(M2:M31,"YES")')
    ws_results.cell(row=37, column=1, value="Total cost (within budget):")
    ws_results.cell(row=37, column=2, value='=SUMIF(M2:M31,"YES",I2:I31)')
    ws_results.cell(row=38, column=1, value="Total impressions (within budget):")
    ws_results.cell(row=38, column=2, value='=SUMIF(M2:M31,"YES",J2:J31)')
    ws_results.cell(row=39, column=1, value="Goal value:")
    ws_results.cell(row=39, column=2, value="='Campaign Inputs'!B9")
    ws_results.cell(row=40, column=1, value="Goal achievement %:")
    ws_results.cell(row=40, column=2, value="=L38/'Campaign Inputs'!B9*100")
    
    for col in range(1, len(result_headers) + 1):
        ws_results.column_dimensions[get_column_letter(col)].width = 15
    ws_results.column_dimensions['C'].width = 30
    
    # Sheet 5: City Allocation
    ws_city = wb.create_sheet("City Allocation")
    
    city_headers = ["City", "Tier", "Min %", "Max %", "Inventory Count", "Score Total", "Allocated %", "Allocated Budget"]
    
    for col, header in enumerate(city_headers, 1):
        cell = ws_city.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    all_cities = []
    for country, cities in CITIES.items():
        for city in cities:
            tier = CITY_TIERS.get(city, 3)
            all_cities.append((city, tier))
    
    for row_idx, (city, tier) in enumerate(all_cities[:20], 2):
        ws_city.cell(row=row_idx, column=1, value=city)
        ws_city.cell(row=row_idx, column=2, value=tier)
        min_pct = 0.15 if tier == 1 else (0.10 if tier == 2 else 0.05)
        max_pct = 0.50 if tier == 1 else (0.40 if tier == 2 else 1.0)
        ws_city.cell(row=row_idx, column=3, value=min_pct)
        ws_city.cell(row=row_idx, column=4, value=max_pct)
        ws_city.cell(row=row_idx, column=3).number_format = '0%'
        ws_city.cell(row=row_idx, column=4).number_format = '0%'
        
        # Inventory Count - count matching inventories from scoring sheet
        ws_city.cell(row=row_idx, column=5, 
            value=f'=COUNTIF(Scoring!C$2:C$201,A{row_idx})')
        
        # Score Total - sum of scores for this city
        ws_city.cell(row=row_idx, column=6,
            value=f'=SUMIF(Scoring!C$2:C$201,A{row_idx},Scoring!O$2:O$201)')
        
        # Allocated % - simplified allocation (proportional to score, with bounds)
        ws_city.cell(row=row_idx, column=7,
            value=f'=IF(E{row_idx}=0,0,MIN(D{row_idx},MAX(C{row_idx},F{row_idx}/SUM($F$2:$F$21))))')
        ws_city.cell(row=row_idx, column=7).number_format = '0%'
        
        # Allocated Budget
        ws_city.cell(row=row_idx, column=8,
            value=f"=G{row_idx}*'Campaign Inputs'!$B$4")
        ws_city.cell(row=row_idx, column=8).number_format = '#,##0'
    
    # Add total row
    ws_city.cell(row=23, column=1, value="TOTAL").font = Font(bold=True)
    ws_city.cell(row=23, column=5, value="=SUM(E2:E21)")
    ws_city.cell(row=23, column=6, value="=SUM(F2:F21)")
    ws_city.cell(row=23, column=7, value="=SUM(G2:G21)")
    ws_city.cell(row=23, column=8, value="=SUM(H2:H21)")
    
    for col in range(1, len(city_headers) + 1):
        ws_city.column_dimensions[get_column_letter(col)].width = 18
    
    # Sheet 6: Format Allocation
    ws_format = wb.create_sheet("Format Allocation")
    
    format_headers = ["Goal Type", "Digital Screen", "Classic", "Transit", "Retail", "Network", "Radio", "Experiential"]
    
    for col, header in enumerate(format_headers, 1):
        cell = ws_format.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    allocations = [
        ("Impressions", 0.35, 0.25, 0.20, 0.10, 0.07, 0.00, 0.03),
        ("Reach", 0.15, 0.30, 0.25, 0.15, 0.10, 0.00, 0.05),
        ("SOV", 0.50, 0.15, 0.15, 0.10, 0.10, 0.00, 0.00),
        ("Ad Plays", 0.50, 0.15, 0.20, 0.10, 0.05, 0.00, 0.00),
        ("Carbon Emission", 0.05, 0.45, 0.30, 0.15, 0.05, 0.00, 0.00),
    ]
    
    for row_idx, row_data in enumerate(allocations, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_format.cell(row=row_idx, column=col_idx, value=value)
            if col_idx > 1:
                cell.number_format = '0%'
    
    for col in range(1, len(format_headers) + 1):
        ws_format.column_dimensions[get_column_letter(col)].width = 15
    
    # Sheet 7: Schedule Output
    ws_schedule = wb.create_sheet("Schedule Output")
    
    schedule_headers = [
        "Inventory ID", "Inventory Name", "City", "Start Date", "End Date", 
        "Days Booked", "Hours/Day", "Total Hours", "Total Plays", 
        "Total Cost", "Est. Impressions", "Est. Reach", "CO2 Produced"
    ]
    
    for col, header in enumerate(schedule_headers, 1):
        cell = ws_schedule.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    ws_schedule.cell(row=2, column=1, value="Schedules are generated based on selection from Scoring sheet")
    ws_schedule.merge_cells('A2:M2')
    
    for col in range(1, len(schedule_headers) + 1):
        ws_schedule.column_dimensions[get_column_letter(col)].width = 15
    ws_schedule.column_dimensions['B'].width = 30
    
    # Sheet 8: Instructions
    ws_help = wb.create_sheet("Instructions")
    
    instructions = [
        ("MW Planner Recommendation Engine V2 - Excel Simulation", "header"),
        ("Version: 1.0 | Created: January 2026", None),
        ("", None),
        ("ABOUT THIS FILE:", "header"),
        ("", None),
        ("This Excel file is a SIMULATION of the MW Planner Recommendation Engine V2.", None),
        ("It demonstrates the core scoring and selection logic with 200 sample inventories.", None),
        ("Use it to understand how different inputs affect inventory selection and scoring.", None),
        ("", None),
        ("SCOPE - WHAT THIS FILE DOES:", "subheader"),
        ("This Excel file simulates SCORING and RANKING only:", None),
        ("  + Calculates 8-factor scores for each inventory", None),
        ("  + Filters by country, city, and venue type preference", None),
        ("  + Ranks inventories by total score", None),
        ("  + Shows which inventories fit within budget", None),
        ("  + Demonstrates how goal type affects measure_fit scoring", None),
        ("", None),
        ("SCOPE - WHAT REQUIRES FULL SOFTWARE:", "subheader"),
        ("These features require the actual MW Planner application:", None),
        ("  - City/format diversity ENFORCEMENT (caps, minimums, reallocation)", None),
        ("  - Partial day booking with minimum day validation", None),
        ("  - CPM-based hour selection within schedules", None),
        ("  - Iterative budget allocation across cities/formats", None),
        ("  - Schedule creation with operating hours validation", None),
        ("The City Allocation sheet shows diversity CALCULATIONS but doesn't", None),
        ("constrain selection. Full diversity enforcement is algorithmic.", None),
        ("", None),
        ("HOW TO USE THIS FILE:", "header"),
        ("", None),
        ("1. Campaign Inputs Sheet (EDITABLE):", "subheader"),
        ("   Yellow cells are editable - change these to test scenarios:", None),
        ("   - Country: Malaysia, Singapore, Indonesia, Thailand, Philippines", None),
        ("   - Target City: Leave blank for country-wide, or enter specific city", None),
        ("   - Budget: Campaign budget in local currency", None),
        ("   - Campaign Dates: Start and end dates (days calculated automatically)", None),
        ("   - Goal Type: Impressions, Reach, SOV, Ad Plays, or Carbon Emission", None),
        ("   - Goal Value: Target number for the goal (e.g., 500000 impressions)", None),
        ("   - Scoring Weights: Adjust to prioritize different factors", None),
        ("", None),
        ("2. Inventory Data Sheet (SAMPLE DATA):", "subheader"),
        ("   200 inventories across 5 countries with realistic metrics:", None),
        ("   - Daily impressions, reach, cost, CO2 per play", None),
        ("   - Operating hours, minimum days, availability percentage", None),
        ("   - Quality, audience match, brand fit scores", None),
        ("   You can modify this data to test edge cases.", None),
        ("", None),
        ("3. Scoring Sheet (AUTO-CALCULATED):", "subheader"),
        ("   Shows scores for each inventory based on Campaign Inputs:", None),
        ("   - Country Match: 1 if inventory country = campaign country", None),
        ("   - City Match: 1 if city matches or no city filter applied", None),
        ("   - measure_fit: How well inventory delivers the goal (0-100)", None),
        ("   - geo_fit: Location quality (Tier 1=100, Tier 2=80, Tier 3=60)", None),
        ("   - availability, audience_fit, brand_fit, quality_fit, time_fit: From data", None),
        ("   - Total Score: Weighted sum of all factors", None),
        ("   - Rank: Position among all inventories (1 = best)", None),
        ("", None),
        ("4. Selection Results Sheet (AUTO-CALCULATED):", "subheader"),
        ("   Shows top 30 inventories by rank with budget analysis:", None),
        ("   - Campaign Cost/Impressions: Daily values × campaign days", None),
        ("   - Cumulative Budget: Running total of selected inventory costs", None),
        ("   - Within Budget: Shows when cumulative exceeds budget", None),
        ("   - Summary section shows total selected within budget", None),
        ("", None),
        ("5. City Allocation Sheet (AUTO-CALCULATED):", "subheader"),
        ("   Geographic diversity analysis:", None),
        ("   - Tier 1 cities (KL, Singapore, etc.): 15% min, 50% max", None),
        ("   - Tier 2 cities (Penang, JB, etc.): 10% min, 40% max", None),
        ("   - Tier 3 cities: 5% min, no max", None),
        ("   - Allocated % shows bounded allocation based on score totals", None),
        ("", None),
        ("6. Format Allocation Sheet (REFERENCE):", "subheader"),
        ("   Default format allocation percentages by goal type:", None),
        ("   - Impressions: Digital 35%, Classic 25%, Transit 20%", None),
        ("   - Carbon Emission: Classic 45%, Transit 30%, Digital 5%", None),
        ("", None),
        ("SCENARIOS TO TEST:", "header"),
        ("", None),
        ("Scenario 1: Change Goal Type", "subheader"),
        ("   1. Go to Campaign Inputs sheet", None),
        ("   2. Change cell B8 from 'Impressions' to 'Carbon Emission'", None),
        ("   3. Check Scoring sheet - measure_fit now rewards low CO2", None),
        ("   4. Check Selection Results - different inventories ranked higher", None),
        ("", None),
        ("Scenario 2: Target Specific City", "subheader"),
        ("   1. In Campaign Inputs, set B3 to 'Kuala Lumpur'", None),
        ("   2. Check Scoring sheet - non-KL inventories have 0 score", None),
        ("   3. Check City Allocation - only KL shows inventory count", None),
        ("", None),
        ("Scenario 3: Reduce Budget", "subheader"),
        ("   1. Change B4 from 100000 to 20000", None),
        ("   2. Check Scoring - expensive inventories have lower budget_fit", None),
        ("   3. Check Selection Results - fewer inventories fit within budget", None),
        ("", None),
        ("Scenario 4: Change Country", "subheader"),
        ("   1. Change B2 from 'Malaysia' to 'Singapore'", None),
        ("   2. All scoring recalculates for Singapore inventories", None),
        ("   3. City Allocation shows Singapore cities", None),
        ("", None),
        ("Scenario 5: Test Venue Type Preference", "subheader"),
        ("   1. In Campaign Inputs, set B10 to 'Transit'", None),
        ("   2. Check Scoring sheet - Venue Match column shows 0 for non-Transit", None),
        ("   3. Non-Transit inventories get Total Score = 0", None),
        ("   4. Selection Results shows only Transit inventories", None),
        ("", None),
        ("Scenario 6: Adjust Scoring Weights", "subheader"),
        ("   1. Change measure_fit weight (B13) from 0.20 to 0.40", None),
        ("   2. Reduce other weights proportionally", None),
        ("   3. Goal delivery becomes more important in total score", None),
        ("", None),
        ("KEY FORMULAS USED:", "header"),
        ("", None),
        ("Total Score = Σ(factor_score × weight)", None),
        ("", None),
        ("measure_fit by goal type:", None),
        ("  Impressions: min(100, daily_imps × days / goal × 100)", None),
        ("  Reach: min(100, daily_reach × days / goal × 100)", None),
        ("  SOV: Fixed at 80 (simplified)", None),
        ("  Ad Plays: min(100, plays_per_hour × hours × days / goal × 100)", None),
        ("  Carbon: 100 - min(100, co2_per_play × 1000) - lower CO2 = higher score", None),
        ("", None),
        ("geo_fit: Tier 1=100, Tier 2=80, Tier 3=60 (when country/city match)", None),
        ("budget_fit: 100 if affordable, penalized proportionally if over budget", None),
        ("", None),
        ("DOCUMENTATION REFERENCE:", "subheader"),
        ("Full documentation: docs/recommendation-engine-v2-planner.md", None),
    ]
    
    for row_idx, (text, style) in enumerate(instructions, 1):
        cell = ws_help.cell(row=row_idx, column=1, value=text)
        if style == "header":
            cell.font = Font(bold=True, size=14)
        elif style == "subheader":
            cell.font = Font(bold=True, size=11)
    
    ws_help.column_dimensions['A'].width = 80
    
    return wb

if __name__ == "__main__":
    print("Generating Excel simulation file...")
    wb = create_workbook()
    output_path = "docs/mw-planner-recommendation-engine-simulation.xlsx"
    wb.save(output_path)
    print(f"Excel file saved to: {output_path}")
    print("Open the file in Excel or Google Sheets to test scenarios.")
