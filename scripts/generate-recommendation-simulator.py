#!/usr/bin/env python3
"""
MW Planner Recommendation Engine V2 - Interactive Simulator
Generates a comprehensive Excel/Google Sheets file with:
- Malaysia and Japan inventory data with city tiers
- Input controls: country, city, venue type, audience targeting
- Scoring formulas (8 factors)
- Budget allocation and selection logic
- Schedule creation with visualization
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta

def create_simulator():
    wb = Workbook()
    
    # Styling
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    input_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    label_font = Font(bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # ========== SHEET 1: INPUTS ==========
    ws_inputs = wb.active
    ws_inputs.title = "1_Inputs"
    
    # Title
    ws_inputs['A1'] = "MW PLANNER - RECOMMENDATION ENGINE V2 SIMULATOR"
    ws_inputs['A1'].font = Font(bold=True, size=16, color="1F4E79")
    ws_inputs.merge_cells('A1:G1')
    
    ws_inputs['A2'] = "Configure campaign parameters. All selections filter and score inventories automatically."
    ws_inputs['A2'].font = Font(italic=True, size=10)
    ws_inputs.merge_cells('A2:G2')
    
    # === CAMPAIGN BASICS ===
    ws_inputs['A4'] = "CAMPAIGN BASICS"
    ws_inputs['A4'].font = header_font
    ws_inputs['A4'].fill = header_fill
    ws_inputs.merge_cells('A4:B4')
    
    # Country (B5)
    ws_inputs['A5'] = "Country *"
    ws_inputs['A5'].font = label_font
    ws_inputs['B5'] = "Malaysia"
    ws_inputs['B5'].fill = input_fill
    ws_inputs['B5'].border = thin_border
    country_dv = DataValidation(type="list", formula1='"Malaysia,Japan"', allow_blank=False)
    ws_inputs.add_data_validation(country_dv)
    country_dv.add(ws_inputs['B5'])
    
    # Currency display (C5) - auto calculated
    ws_inputs['C5'] = '=IF(B5="Malaysia","MYR","JPY")'
    ws_inputs['C5'].font = Font(bold=True, color="1F4E79")
    
    # City (B6)
    ws_inputs['A6'] = "City Filter"
    ws_inputs['A6'].font = label_font
    ws_inputs['B6'] = "All Cities"
    ws_inputs['B6'].fill = input_fill
    ws_inputs['B6'].border = thin_border
    # Dynamic city list based on country
    city_dv = DataValidation(type="list", formula1='"All Cities,Kuala Lumpur,KLIA,Petaling Jaya,Penang,Johor Bahru,Ipoh,Melaka,Kuching,Kota Kinabalu,Tokyo,Osaka,Nagoya,Yokohama,Kyoto,Fukuoka"', allow_blank=False)
    ws_inputs.add_data_validation(city_dv)
    city_dv.add(ws_inputs['B6'])
    
    # Start Date (B7)
    ws_inputs['A7'] = "Start Date *"
    ws_inputs['A7'].font = label_font
    ws_inputs['B7'] = "2026-02-01"
    ws_inputs['B7'].fill = input_fill
    ws_inputs['B7'].border = thin_border
    
    # End Date (B8)
    ws_inputs['A8'] = "End Date *"
    ws_inputs['A8'].font = label_font
    ws_inputs['B8'] = "2026-02-28"
    ws_inputs['B8'].fill = input_fill
    ws_inputs['B8'].border = thin_border
    
    # Budget (B9)
    ws_inputs['A9'] = "Budget *"
    ws_inputs['A9'].font = label_font
    ws_inputs['B9'] = 100000
    ws_inputs['B9'].fill = input_fill
    ws_inputs['B9'].border = thin_border
    ws_inputs['B9'].number_format = '#,##0'
    ws_inputs['C9'] = '=C5'  # Currency label
    
    # Goal Type (B10)
    ws_inputs['A10'] = "Goal Type *"
    ws_inputs['A10'].font = label_font
    ws_inputs['B10'] = "Impressions"
    ws_inputs['B10'].fill = input_fill
    ws_inputs['B10'].border = thin_border
    goal_dv = DataValidation(type="list", formula1='"Impressions,Reach,Ad Plays,SOV,Carbon Emission"', allow_blank=False)
    ws_inputs.add_data_validation(goal_dv)
    goal_dv.add(ws_inputs['B10'])
    
    # Goal Value (B11)
    ws_inputs['A11'] = "Goal Value"
    ws_inputs['A11'].font = label_font
    ws_inputs['B11'] = 1000000
    ws_inputs['B11'].fill = input_fill
    ws_inputs['B11'].border = thin_border
    ws_inputs['B11'].number_format = '#,##0'
    
    # === VENUE TARGETING ===
    ws_inputs['A13'] = "VENUE TARGETING"
    ws_inputs['A13'].font = header_font
    ws_inputs['A13'].fill = header_fill
    ws_inputs.merge_cells('A13:B13')
    
    # Venue Type Filter (B14)
    ws_inputs['A14'] = "Venue Type"
    ws_inputs['A14'].font = label_font
    ws_inputs['B14'] = "All Venues"
    ws_inputs['B14'].fill = input_fill
    ws_inputs['B14'].border = thin_border
    venue_dv = DataValidation(type="list", formula1='"All Venues,Transit: Airport,Transit: Rail/Metro Station,Transit: Bus,Retail: Shopping Center,Retail: High Street,Billboard: Highway,Billboard: Roadside,Office Building,Audio: Radio"', allow_blank=False)
    ws_inputs.add_data_validation(venue_dv)
    venue_dv.add(ws_inputs['B14'])
    
    # Classification Filter (B15)
    ws_inputs['A15'] = "Classification"
    ws_inputs['A15'].font = label_font
    ws_inputs['B15'] = "All"
    ws_inputs['B15'].fill = input_fill
    ws_inputs['B15'].border = thin_border
    class_dv = DataValidation(type="list", formula1='"All,Digital,Classic,Audio"', allow_blank=False)
    ws_inputs.add_data_validation(class_dv)
    class_dv.add(ws_inputs['B15'])
    
    # Type Filter (B16)
    ws_inputs['A16'] = "Type"
    ws_inputs['A16'].font = label_font
    ws_inputs['B16'] = "All"
    ws_inputs['B16'].fill = input_fill
    ws_inputs['B16'].border = thin_border
    type_dv = DataValidation(type="list", formula1='"All,OOH,Transit,Retail,Network,Radio"', allow_blank=False)
    ws_inputs.add_data_validation(type_dv)
    type_dv.add(ws_inputs['B16'])
    
    # === AUDIENCE TARGETING (matches campaign creation flow) ===
    ws_inputs['D4'] = "AUDIENCE TARGETING"
    ws_inputs['D4'].font = header_font
    ws_inputs['D4'].fill = header_fill
    ws_inputs.merge_cells('D4:E4')
    
    # Demographics
    ws_inputs['D5'] = "Age Group"
    ws_inputs['D5'].font = label_font
    ws_inputs['E5'] = "25-44"
    ws_inputs['E5'].fill = input_fill
    ws_inputs['E5'].border = thin_border
    age_dv = DataValidation(type="list", formula1='"All Ages,18-24,25-34,35-44,45-54,55+,25-44"', allow_blank=False)
    ws_inputs.add_data_validation(age_dv)
    age_dv.add(ws_inputs['E5'])
    
    ws_inputs['D6'] = "Gender"
    ws_inputs['D6'].font = label_font
    ws_inputs['E6'] = "All"
    ws_inputs['E6'].fill = input_fill
    ws_inputs['E6'].border = thin_border
    gender_dv = DataValidation(type="list", formula1='"All,Male,Female"', allow_blank=False)
    ws_inputs.add_data_validation(gender_dv)
    gender_dv.add(ws_inputs['E6'])
    
    # Income
    ws_inputs['D7'] = "Income Level"
    ws_inputs['D7'].font = label_font
    ws_inputs['E7'] = "All"
    ws_inputs['E7'].fill = input_fill
    ws_inputs['E7'].border = thin_border
    income_dv = DataValidation(type="list", formula1='"All,Low,Middle,High,Affluent"', allow_blank=False)
    ws_inputs.add_data_validation(income_dv)
    income_dv.add(ws_inputs['E7'])
    
    # Interests
    ws_inputs['D8'] = "Interest"
    ws_inputs['D8'].font = label_font
    ws_inputs['E8'] = "All"
    ws_inputs['E8'].fill = input_fill
    ws_inputs['E8'].border = thin_border
    interest_dv = DataValidation(type="list", formula1='"All,Travel,Fashion,Technology,Food & Beverage,Sports,Finance,Entertainment,Automotive"', allow_blank=False)
    ws_inputs.add_data_validation(interest_dv)
    interest_dv.add(ws_inputs['E8'])
    
    # Behaviors
    ws_inputs['D9'] = "Behavior"
    ws_inputs['D9'].font = label_font
    ws_inputs['E9'] = "All"
    ws_inputs['E9'].fill = input_fill
    ws_inputs['E9'].border = thin_border
    behavior_dv = DataValidation(type="list", formula1='"All,Commuters,Shoppers,Business Travelers,Tourists,Residents,Frequent Flyers"', allow_blank=False)
    ws_inputs.add_data_validation(behavior_dv)
    behavior_dv.add(ws_inputs['E9'])
    
    # === CALCULATED VALUES ===
    ws_inputs['D11'] = "CALCULATED VALUES"
    ws_inputs['D11'].font = header_font
    ws_inputs['D11'].fill = header_fill
    ws_inputs.merge_cells('D11:F11')
    
    calc_values = [
        ("D12", "Campaign Days:", "E12", '=DATEDIF(B7,B8,"D")+1', "F12", "days"),
        ("D13", "Pricing Model:", "E13", '=IF(OR(B10="Impressions",B10="Reach",B10="Carbon Emission"),"CPM","CPS")', "F13", ""),
        ("D14", "Currency:", "E14", '=C5', "F14", ""),
    ]
    
    for label_cell, label, val_cell, formula, unit_cell, unit in calc_values:
        ws_inputs[label_cell] = label
        ws_inputs[label_cell].font = label_font
        ws_inputs[val_cell] = formula
        ws_inputs[unit_cell] = unit
    
    # === BUDGET ALLOCATION ===
    ws_inputs['D16'] = "BUDGET ALLOCATION (auto by goal)"
    ws_inputs['D16'].font = header_font
    ws_inputs['D16'].fill = header_fill
    ws_inputs.merge_cells('D16:G16')
    
    ws_inputs['D17'] = "Classification"
    ws_inputs['E17'] = "Allocation %"
    ws_inputs['F17'] = "Amount"
    ws_inputs['G17'] = "Rationale"
    for col in ['D', 'E', 'F', 'G']:
        ws_inputs[f'{col}17'].font = Font(bold=True)
        ws_inputs[f'{col}17'].fill = section_fill
    
    # Digital
    ws_inputs['D18'] = "Digital"
    ws_inputs['E18'] = '=IF(B10="Impressions",0.60,IF(B10="Reach",0.55,IF(B10="Ad Plays",0.95,IF(B10="SOV",0.95,0.20))))'
    ws_inputs['E18'].number_format = '0%'
    ws_inputs['F18'] = '=E18*B9'
    ws_inputs['F18'].number_format = '#,##0'
    ws_inputs['G18'] = '=IF(B10="Ad Plays","Ad plays require digital screens",IF(B10="SOV","SOV requires slot-based inventory","Balanced for goal type"))'
    
    # Classic
    ws_inputs['D19'] = "Classic"
    ws_inputs['E19'] = '=IF(B10="Impressions",0.35,IF(B10="Reach",0.40,IF(B10="Ad Plays",0,IF(B10="SOV",0,0.75))))'
    ws_inputs['E19'].number_format = '0%'
    ws_inputs['F19'] = '=E19*B9'
    ws_inputs['F19'].number_format = '#,##0'
    ws_inputs['G19'] = '=IF(OR(B10="Ad Plays",B10="SOV"),"Classic cannot produce plays/slots","Geographic coverage")'
    
    # Audio
    ws_inputs['D20'] = "Audio"
    ws_inputs['E20'] = '=1-E18-E19'
    ws_inputs['E20'].number_format = '0%'
    ws_inputs['F20'] = '=E20*B9'
    ws_inputs['F20'].number_format = '#,##0'
    ws_inputs['G20'] = "Remaining allocation"
    
    # Column widths
    ws_inputs.column_dimensions['A'].width = 18
    ws_inputs.column_dimensions['B'].width = 18
    ws_inputs.column_dimensions['C'].width = 8
    ws_inputs.column_dimensions['D'].width = 18
    ws_inputs.column_dimensions['E'].width = 16
    ws_inputs.column_dimensions['F'].width = 14
    ws_inputs.column_dimensions['G'].width = 35
    
    # ========== SHEET 2: INVENTORY DATA ==========
    ws_inv = wb.create_sheet("2_Inventory_Data")
    
    # Headers - includes City Tier
    inv_headers = [
        "ID", "Name", "Country", "City", "City Tier", "Classification", "Type", "Format",
        "Venue Type (IAB)", "CPM", "CPS", "Daily Impressions", "Daily Reach",
        "Operating Hours", "Slots Per Loop", "Slot Duration (sec)", "CO2 per Play (kg)",
        "Age 18-24", "Age 25-34", "Age 35-44", "Age 45-54", "Age 55+",
        "Male %", "Female %", "Income Low", "Income Middle", "Income High", "Income Affluent",
        "Interest Travel", "Interest Fashion", "Interest Tech", "Interest Food",
        "Behavior Commuters", "Behavior Shoppers", "Behavior Business", "Behavior Tourists", "Behavior Residents"
    ]
    
    for col, header in enumerate(inv_headers, 1):
        cell = ws_inv.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Malaysia Inventory Data (20 inventories) - with City Tier
    # CPM rates reduced for demo budget of 100,000 MYR over 28 days
    malaysia_data = [
        # ID, Name, Country, City, Tier, Class, Type, Format, Venue, CPM, CPS, DailyImp, DailyReach, OpHrs, Slots, SlotDur, CO2, Age profiles, Gender, Income, Interests, Behaviors
        ["MY001", "KLCC Suria Digital Screen", "Malaysia", "Kuala Lumpur", 1, "Digital", "Retail", "Mall Digital Screen", "Retail: Shopping Center", 8, 3, 85000, 42000, 16, 6, 10, 0.02, 0.25, 0.35, 0.25, 0.10, 0.05, 0.45, 0.55, 0.15, 0.50, 0.30, 0.05, 0.40, 0.35, 0.20, 0.30, 0.15, 0.60, 0.10, 0.25, 0.20],
        ["MY002", "KLIA Terminal 1 Arrival", "Malaysia", "KLIA", 1, "Digital", "Transit", "Airport Digital", "Transit: Airport", 12, 4, 120000, 95000, 20, 4, 15, 0.03, 0.20, 0.30, 0.30, 0.15, 0.05, 0.55, 0.45, 0.10, 0.35, 0.45, 0.10, 0.70, 0.20, 0.15, 0.20, 0.05, 0.15, 0.40, 0.80, 0.05],
        ["MY003", "Federal Highway Billboard", "Malaysia", "Petaling Jaya", 2, "Digital", "OOH", "Digital Billboard", "Billboard: Highway", 5, 2, 150000, 120000, 18, 4, 15, 0.04, 0.20, 0.35, 0.30, 0.10, 0.05, 0.60, 0.40, 0.20, 0.50, 0.25, 0.05, 0.30, 0.25, 0.25, 0.25, 0.70, 0.10, 0.30, 0.05, 0.40],
        ["MY004", "Pavilion KL Entrance", "Malaysia", "Kuala Lumpur", 1, "Digital", "Retail", "Mall Digital Screen", "Retail: Shopping Center", 9, 3, 95000, 48000, 14, 6, 10, 0.02, 0.30, 0.35, 0.20, 0.10, 0.05, 0.40, 0.60, 0.10, 0.40, 0.40, 0.10, 0.35, 0.50, 0.20, 0.35, 0.10, 0.70, 0.15, 0.20, 0.15],
        ["MY005", "KL Sentral LRT Station", "Malaysia", "Kuala Lumpur", 1, "Digital", "Transit", "Station Digital", "Transit: Rail/Metro Station", 6, 2, 180000, 85000, 18, 6, 10, 0.025, 0.25, 0.40, 0.25, 0.08, 0.02, 0.52, 0.48, 0.25, 0.55, 0.18, 0.02, 0.35, 0.20, 0.30, 0.25, 0.80, 0.05, 0.35, 0.10, 0.30],
        ["MY006", "Menara KEN Office Lobby", "Malaysia", "Kuala Lumpur", 1, "Digital", "Network", "Office Lobby Network", "Office Building", 5, 1.5, 25000, 8000, 12, 4, 10, 0.015, 0.15, 0.45, 0.30, 0.08, 0.02, 0.55, 0.45, 0.05, 0.40, 0.45, 0.10, 0.25, 0.30, 0.50, 0.20, 0.10, 0.05, 0.90, 0.02, 0.05],
        ["MY007", "Bangsar LRT Station", "Malaysia", "Kuala Lumpur", 1, "Digital", "Transit", "Station Digital", "Transit: Rail/Metro Station", 6, 2, 65000, 32000, 16, 6, 10, 0.02, 0.30, 0.40, 0.20, 0.07, 0.03, 0.48, 0.52, 0.15, 0.55, 0.25, 0.05, 0.40, 0.35, 0.25, 0.35, 0.75, 0.15, 0.20, 0.08, 0.25],
        ["MY008", "Jalan Bukit Bintang LED", "Malaysia", "Kuala Lumpur", 1, "Digital", "OOH", "LED Screen", "Billboard: Roadside", 10, 3.5, 200000, 150000, 24, 4, 15, 0.05, 0.30, 0.35, 0.20, 0.10, 0.05, 0.50, 0.50, 0.20, 0.45, 0.30, 0.05, 0.45, 0.40, 0.25, 0.40, 0.30, 0.50, 0.15, 0.40, 0.15],
        ["MY009", "Penang Gurney Plaza", "Malaysia", "Penang", 2, "Digital", "Retail", "Mall Digital Screen", "Retail: Shopping Center", 7, 2.5, 55000, 28000, 14, 6, 10, 0.02, 0.25, 0.30, 0.25, 0.12, 0.08, 0.45, 0.55, 0.20, 0.55, 0.22, 0.03, 0.35, 0.30, 0.20, 0.35, 0.20, 0.65, 0.10, 0.15, 0.25],
        ["MY010", "Johor Bahru City Square", "Malaysia", "Johor Bahru", 2, "Digital", "Retail", "Mall Digital Screen", "Retail: Shopping Center", 6, 2, 45000, 22000, 12, 6, 10, 0.018, 0.30, 0.35, 0.20, 0.10, 0.05, 0.48, 0.52, 0.25, 0.55, 0.18, 0.02, 0.30, 0.25, 0.20, 0.30, 0.25, 0.60, 0.10, 0.20, 0.30],
        ["MY011", "Kota Kinabalu Airport", "Malaysia", "Kota Kinabalu", 2, "Digital", "Transit", "Airport Digital", "Transit: Airport", 8, 3, 35000, 28000, 16, 4, 15, 0.025, 0.25, 0.35, 0.25, 0.10, 0.05, 0.52, 0.48, 0.15, 0.45, 0.35, 0.05, 0.60, 0.20, 0.15, 0.25, 0.10, 0.15, 0.35, 0.70, 0.10],
        ["MY012", "Sprint Highway Billboard", "Malaysia", "Petaling Jaya", 2, "Classic", "OOH", "Bulletin", "Billboard: Highway", 4, 0, 130000, 105000, 24, 0, 0, 0, 0.20, 0.35, 0.30, 0.10, 0.05, 0.60, 0.40, 0.25, 0.50, 0.22, 0.03, 0.25, 0.20, 0.25, 0.25, 0.70, 0.10, 0.25, 0.05, 0.35],
        ["MY013", "LDP Highway 48 Sheet", "Malaysia", "Petaling Jaya", 2, "Classic", "OOH", "48 Sheet", "Billboard: Highway", 3, 0, 95000, 80000, 24, 0, 0, 0, 0.22, 0.35, 0.28, 0.10, 0.05, 0.62, 0.38, 0.30, 0.50, 0.18, 0.02, 0.20, 0.18, 0.22, 0.22, 0.72, 0.08, 0.22, 0.04, 0.40],
        ["MY014", "MRR2 Wallscape", "Malaysia", "Kuala Lumpur", 1, "Classic", "OOH", "Wallscape", "Billboard: Roadside", 5, 0, 110000, 90000, 24, 0, 0, 0, 0.20, 0.32, 0.30, 0.12, 0.06, 0.58, 0.42, 0.28, 0.48, 0.20, 0.04, 0.22, 0.20, 0.25, 0.28, 0.68, 0.12, 0.28, 0.06, 0.30],
        ["MY015", "Rapid KL Bus Wrap", "Malaysia", "Kuala Lumpur", 1, "Classic", "Transit", "Bus Exterior", "Transit: Bus", 3.5, 0, 75000, 60000, 16, 0, 0, 0, 0.28, 0.38, 0.22, 0.08, 0.04, 0.50, 0.50, 0.35, 0.50, 0.13, 0.02, 0.25, 0.22, 0.20, 0.28, 0.65, 0.25, 0.15, 0.10, 0.35],
        ["MY016", "Ipoh Parade Mall", "Malaysia", "Ipoh", 2, "Digital", "Retail", "Mall Digital Screen", "Retail: Shopping Center", 5, 1.5, 28000, 14000, 12, 6, 10, 0.015, 0.22, 0.28, 0.28, 0.15, 0.07, 0.45, 0.55, 0.30, 0.50, 0.18, 0.02, 0.25, 0.25, 0.18, 0.32, 0.25, 0.55, 0.12, 0.12, 0.35],
        ["MY017", "Melaka Dataran Pahlawan", "Malaysia", "Melaka", 3, "Digital", "Retail", "Mall Digital Screen", "Retail: Shopping Center", 5, 1.5, 32000, 16000, 12, 6, 10, 0.018, 0.25, 0.30, 0.25, 0.12, 0.08, 0.48, 0.52, 0.28, 0.52, 0.18, 0.02, 0.35, 0.28, 0.18, 0.30, 0.22, 0.55, 0.10, 0.25, 0.30],
        ["MY018", "Kuching Airport", "Malaysia", "Kuching", 2, "Digital", "Transit", "Airport Digital", "Transit: Airport", 7, 2.5, 22000, 18000, 14, 4, 15, 0.022, 0.22, 0.32, 0.28, 0.12, 0.06, 0.54, 0.46, 0.18, 0.48, 0.30, 0.04, 0.55, 0.22, 0.18, 0.25, 0.12, 0.18, 0.38, 0.65, 0.12],
        ["MY019", "BFM 89.9 Radio", "Malaysia", "Kuala Lumpur", 1, "Audio", "Radio", "AM/FM Radio Spot", "Audio: Radio", 1.5, 0.5, 350000, 180000, 24, 0, 30, 0.001, 0.18, 0.35, 0.30, 0.12, 0.05, 0.52, 0.48, 0.10, 0.45, 0.40, 0.05, 0.30, 0.25, 0.35, 0.30, 0.55, 0.20, 0.45, 0.08, 0.20],
        ["MY020", "Hitz FM Radio", "Malaysia", "Kuala Lumpur", 1, "Audio", "Radio", "AM/FM Radio Spot", "Audio: Radio", 2, 0.8, 450000, 220000, 24, 0, 30, 0.001, 0.35, 0.35, 0.18, 0.08, 0.04, 0.48, 0.52, 0.25, 0.50, 0.22, 0.03, 0.28, 0.32, 0.22, 0.38, 0.45, 0.35, 0.25, 0.12, 0.25],
    ]
    
    # Japan Inventory Data (20 inventories) - with City Tier
    # CPM rates reduced for demo budget compatibility (Japan uses JPY, rates scaled for demo)
    japan_data = [
        ["JP001", "Shibuya Crossing LED", "Japan", "Tokyo", 1, "Digital", "OOH", "LED Screen", "Billboard: Roadside", 50, 18, 500000, 380000, 24, 4, 15, 0.06, 0.35, 0.35, 0.18, 0.08, 0.04, 0.52, 0.48, 0.15, 0.45, 0.32, 0.08, 0.45, 0.50, 0.35, 0.40, 0.30, 0.45, 0.20, 0.55, 0.10],
        ["JP002", "Shinjuku Station Digital", "Japan", "Tokyo", 1, "Digital", "Transit", "Station Digital", "Transit: Rail/Metro Station", 35, 12, 650000, 320000, 20, 6, 10, 0.04, 0.30, 0.40, 0.20, 0.07, 0.03, 0.50, 0.50, 0.20, 0.55, 0.22, 0.03, 0.35, 0.40, 0.30, 0.30, 0.75, 0.15, 0.35, 0.15, 0.20],
        ["JP003", "Narita Airport T1", "Japan", "Tokyo", 1, "Digital", "Transit", "Airport Digital", "Transit: Airport", 55, 20, 180000, 150000, 20, 4, 15, 0.045, 0.22, 0.32, 0.28, 0.12, 0.06, 0.55, 0.45, 0.08, 0.35, 0.45, 0.12, 0.72, 0.25, 0.18, 0.22, 0.08, 0.12, 0.42, 0.82, 0.05],
        ["JP004", "Ginza Sony Building", "Japan", "Tokyo", 1, "Digital", "Retail", "Mall Digital Screen", "Retail: Shopping Center", 42, 15, 85000, 45000, 14, 6, 10, 0.035, 0.25, 0.35, 0.25, 0.10, 0.05, 0.45, 0.55, 0.05, 0.35, 0.45, 0.15, 0.40, 0.55, 0.45, 0.35, 0.15, 0.65, 0.25, 0.35, 0.10],
        ["JP005", "Osaka Dotonbori Screen", "Japan", "Osaka", 1, "Digital", "OOH", "LED Screen", "Billboard: Roadside", 40, 14, 350000, 260000, 24, 4, 15, 0.055, 0.32, 0.35, 0.20, 0.08, 0.05, 0.50, 0.50, 0.18, 0.48, 0.28, 0.06, 0.42, 0.45, 0.30, 0.45, 0.28, 0.48, 0.18, 0.48, 0.12],
        ["JP006", "Umeda Station Network", "Japan", "Osaka", 1, "Digital", "Network", "Station Digital", "Transit: Rail/Metro Station", 28, 10, 420000, 210000, 18, 6, 10, 0.038, 0.28, 0.42, 0.20, 0.07, 0.03, 0.52, 0.48, 0.22, 0.55, 0.20, 0.03, 0.32, 0.38, 0.32, 0.28, 0.78, 0.12, 0.38, 0.12, 0.18],
        ["JP007", "Kansai Airport T2", "Japan", "Osaka", 1, "Digital", "Transit", "Airport Digital", "Transit: Airport", 48, 16, 95000, 78000, 18, 4, 15, 0.042, 0.20, 0.30, 0.30, 0.14, 0.06, 0.56, 0.44, 0.10, 0.38, 0.42, 0.10, 0.68, 0.22, 0.15, 0.20, 0.10, 0.15, 0.40, 0.78, 0.05],
        ["JP008", "Harajuku Takeshita St", "Japan", "Tokyo", 1, "Digital", "Retail", "LED Screen", "Retail: High Street", 32, 11, 120000, 70000, 14, 6, 10, 0.032, 0.45, 0.35, 0.12, 0.05, 0.03, 0.35, 0.65, 0.25, 0.50, 0.22, 0.03, 0.35, 0.65, 0.25, 0.45, 0.20, 0.70, 0.08, 0.35, 0.15],
        ["JP009", "Nagoya Station Central", "Japan", "Nagoya", 2, "Digital", "Transit", "Station Digital", "Transit: Rail/Metro Station", 25, 9, 280000, 140000, 18, 6, 10, 0.035, 0.25, 0.40, 0.22, 0.09, 0.04, 0.52, 0.48, 0.22, 0.55, 0.20, 0.03, 0.30, 0.35, 0.32, 0.28, 0.72, 0.15, 0.38, 0.12, 0.22],
        ["JP010", "Kyoto Station Karasuma", "Japan", "Kyoto", 2, "Digital", "Transit", "Station Digital", "Transit: Rail/Metro Station", 22, 8, 180000, 90000, 16, 6, 10, 0.03, 0.22, 0.32, 0.25, 0.14, 0.07, 0.48, 0.52, 0.18, 0.52, 0.26, 0.04, 0.55, 0.30, 0.25, 0.32, 0.60, 0.20, 0.25, 0.45, 0.18],
        ["JP011", "Fukuoka Tenjin Core", "Japan", "Fukuoka", 2, "Digital", "Retail", "Mall Digital Screen", "Retail: Shopping Center", 18, 6, 75000, 38000, 14, 6, 10, 0.028, 0.28, 0.35, 0.22, 0.10, 0.05, 0.45, 0.55, 0.22, 0.52, 0.23, 0.03, 0.35, 0.40, 0.25, 0.38, 0.22, 0.60, 0.15, 0.22, 0.25],
        ["JP012", "Yokohama Bay Quarter", "Japan", "Yokohama", 2, "Digital", "Retail", "Mall Digital Screen", "Retail: Shopping Center", 20, 7, 68000, 35000, 14, 6, 10, 0.025, 0.25, 0.38, 0.22, 0.10, 0.05, 0.48, 0.52, 0.18, 0.50, 0.28, 0.04, 0.38, 0.42, 0.28, 0.35, 0.20, 0.58, 0.18, 0.28, 0.22],
        ["JP013", "Roppongi Hills Tower", "Japan", "Tokyo", 1, "Digital", "Network", "Office Lobby Network", "Office Building", 26, 9, 35000, 12000, 12, 4, 10, 0.02, 0.12, 0.42, 0.32, 0.10, 0.04, 0.58, 0.42, 0.03, 0.35, 0.50, 0.12, 0.35, 0.35, 0.55, 0.25, 0.08, 0.05, 0.88, 0.05, 0.05],
        ["JP014", "Akihabara Electric Town", "Japan", "Tokyo", 1, "Digital", "OOH", "LED Screen", "Retail: High Street", 30, 10, 180000, 100000, 18, 4, 15, 0.04, 0.40, 0.38, 0.15, 0.05, 0.02, 0.70, 0.30, 0.25, 0.50, 0.22, 0.03, 0.28, 0.25, 0.70, 0.32, 0.25, 0.55, 0.15, 0.30, 0.20],
        ["JP015", "Haneda Airport T3", "Japan", "Tokyo", 1, "Digital", "Transit", "Airport Digital", "Transit: Airport", 52, 18, 220000, 185000, 22, 4, 15, 0.048, 0.20, 0.30, 0.30, 0.14, 0.06, 0.54, 0.46, 0.08, 0.32, 0.48, 0.12, 0.75, 0.22, 0.18, 0.20, 0.05, 0.10, 0.45, 0.85, 0.05],
        ["JP016", "Tokyo Metro Marunouchi", "Japan", "Tokyo", 1, "Classic", "Transit", "Station Poster", "Transit: Rail/Metro Station", 12, 0, 380000, 190000, 20, 0, 0, 0, 0.28, 0.42, 0.20, 0.07, 0.03, 0.50, 0.50, 0.22, 0.55, 0.20, 0.03, 0.32, 0.35, 0.30, 0.28, 0.78, 0.12, 0.38, 0.12, 0.18],
        ["JP017", "Tomei Expressway Billboard", "Japan", "Tokyo", 1, "Classic", "OOH", "Bulletin", "Billboard: Highway", 15, 0, 250000, 200000, 24, 0, 0, 0, 0.20, 0.35, 0.30, 0.10, 0.05, 0.62, 0.38, 0.18, 0.48, 0.28, 0.06, 0.28, 0.22, 0.28, 0.25, 0.72, 0.10, 0.32, 0.05, 0.30],
        ["JP018", "Osaka Loop Train Wrap", "Japan", "Osaka", 1, "Classic", "Transit", "Train Wrap", "Transit: Rail/Metro Station", 22, 0, 450000, 280000, 18, 0, 0, 0, 0.30, 0.38, 0.20, 0.08, 0.04, 0.50, 0.50, 0.25, 0.52, 0.20, 0.03, 0.35, 0.38, 0.28, 0.32, 0.72, 0.18, 0.28, 0.15, 0.20],
        ["JP019", "J-Wave 81.3FM", "Japan", "Tokyo", 1, "Audio", "Radio", "AM/FM Radio Spot", "Audio: Radio", 7, 3, 800000, 400000, 24, 0, 30, 0.002, 0.22, 0.38, 0.25, 0.10, 0.05, 0.50, 0.50, 0.12, 0.45, 0.38, 0.05, 0.32, 0.40, 0.35, 0.32, 0.50, 0.25, 0.40, 0.15, 0.18],
        ["JP020", "FM Yokohama 84.7", "Japan", "Yokohama", 2, "Audio", "Radio", "AM/FM Radio Spot", "Audio: Radio", 5.5, 2, 550000, 280000, 24, 0, 30, 0.002, 0.25, 0.35, 0.25, 0.10, 0.05, 0.48, 0.52, 0.18, 0.50, 0.28, 0.04, 0.35, 0.38, 0.30, 0.35, 0.52, 0.28, 0.35, 0.18, 0.22],
    ]
    
    # Write data
    for row_idx, row_data in enumerate(malaysia_data + japan_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws_inv.cell(row=row_idx, column=col_idx, value=value)
    
    # Freeze header row
    ws_inv.freeze_panes = 'A2'
    
    # Column widths
    for col in range(1, len(inv_headers) + 1):
        ws_inv.column_dimensions[get_column_letter(col)].width = 14
    ws_inv.column_dimensions['B'].width = 28
    ws_inv.column_dimensions['I'].width = 25
    
    # ========== SHEET 3: SCORING ==========
    ws_score = wb.create_sheet("3_Scoring")
    
    # Title
    ws_score['A1'] = "INVENTORY SCORING - 8 FACTORS (with filters)"
    ws_score['A1'].font = Font(bold=True, size=14, color="1F4E79")
    ws_score.merge_cells('A1:N1')
    
    ws_score['A2'] = "Scores are calculated only for inventories matching Country, City, Venue, and Classification filters. Non-matching rows show 0."
    ws_score['A2'].font = Font(italic=True, size=9)
    ws_score.merge_cells('A2:N2')
    
    # Scoring Headers
    score_headers = [
        "ID", "Name", "City", "Tier", "Filter Match", "Country", "Geo Fit", "Availability", "Budget Fit",
        "Audience Fit", "Brand Fit", "Quality", "Time Fit", "FINAL SCORE", "RANK"
    ]
    
    for col, header in enumerate(score_headers, 1):
        cell = ws_score.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Scoring formulas for each inventory (rows 5-44)
    for row in range(5, 45):
        inv_row = row - 3  # Inventory data row
        
        # ID
        ws_score.cell(row=row, column=1, value=f"='2_Inventory_Data'!A{inv_row}")
        
        # Name
        ws_score.cell(row=row, column=2, value=f"='2_Inventory_Data'!B{inv_row}")
        
        # City
        ws_score.cell(row=row, column=3, value=f"='2_Inventory_Data'!D{inv_row}")
        
        # Tier
        ws_score.cell(row=row, column=4, value=f"='2_Inventory_Data'!E{inv_row}")
        
        # Filter Match (checks all filters: country, city, venue, classification, type)
        filter_formula = f'''=IF(AND(
'2_Inventory_Data'!C{inv_row}='1_Inputs'!$B$5,
OR('1_Inputs'!$B$6="All Cities",'2_Inventory_Data'!D{inv_row}='1_Inputs'!$B$6),
OR('1_Inputs'!$B$14="All Venues",'2_Inventory_Data'!I{inv_row}='1_Inputs'!$B$14),
OR('1_Inputs'!$B$15="All",'2_Inventory_Data'!F{inv_row}='1_Inputs'!$B$15),
OR('1_Inputs'!$B$16="All",'2_Inventory_Data'!G{inv_row}='1_Inputs'!$B$16)
),"MATCH","FILTERED")'''
        ws_score.cell(row=row, column=5, value=filter_formula.replace('\n', ''))
        
        # Country Match (100 if match, 0 if not)
        ws_score.cell(row=row, column=6, value=f'=IF(E{row}="MATCH",100,0)')
        
        # Geo Fit (based on city tier - Tier 1 = 100, Tier 2 = 85, Tier 3 = 70)
        ws_score.cell(row=row, column=7, value=f'=IF(E{row}="MATCH",IF(D{row}=1,100,IF(D{row}=2,85,70)),0)')
        
        # Availability (assume 80-100%)
        ws_score.cell(row=row, column=8, value=f'=IF(E{row}="MATCH",80+MOD(ROW(),21),0)')
        
        # Budget Fit (based on CPM efficiency - lower CPM = higher score)
        ws_score.cell(row=row, column=9, value=f'=IF(E{row}="MATCH",MAX(0,100-\'2_Inventory_Data\'!J{inv_row}*0.3),0)')
        
        # Audience Fit (complex formula based on selected audience segments)
        audience_formula = f'''=IF(E{row}="FILTERED",0,
ROUND((
IF('1_Inputs'!$E$5="All Ages",50,
IF('1_Inputs'!$E$5="25-44",('2_Inventory_Data'!S{inv_row}+'2_Inventory_Data'!T{inv_row})*100,
IF('1_Inputs'!$E$5="18-24",'2_Inventory_Data'!R{inv_row}*100,
IF('1_Inputs'!$E$5="25-34",'2_Inventory_Data'!S{inv_row}*100,
IF('1_Inputs'!$E$5="35-44",'2_Inventory_Data'!T{inv_row}*100,
IF('1_Inputs'!$E$5="45-54",'2_Inventory_Data'!U{inv_row}*100,
IF('1_Inputs'!$E$5="55+",'2_Inventory_Data'!V{inv_row}*100,50)))))))
+IF('1_Inputs'!$E$8="All",50,
IF('1_Inputs'!$E$8="Travel",'2_Inventory_Data'!AC{inv_row}*100,
IF('1_Inputs'!$E$8="Fashion",'2_Inventory_Data'!AD{inv_row}*100,
IF('1_Inputs'!$E$8="Technology",'2_Inventory_Data'!AE{inv_row}*100,
IF('1_Inputs'!$E$8="Food & Beverage",'2_Inventory_Data'!AF{inv_row}*100,50)))))
+IF('1_Inputs'!$E$9="All",50,
IF('1_Inputs'!$E$9="Commuters",'2_Inventory_Data'!AG{inv_row}*100,
IF('1_Inputs'!$E$9="Shoppers",'2_Inventory_Data'!AH{inv_row}*100,
IF('1_Inputs'!$E$9="Business Travelers",'2_Inventory_Data'!AI{inv_row}*100,
IF('1_Inputs'!$E$9="Tourists",'2_Inventory_Data'!AJ{inv_row}*100,50)))))
)/3,1))'''
        ws_score.cell(row=row, column=10, value=audience_formula.replace('\n', ''))
        
        # Brand Fit (travel brand matches airports/transit)
        ws_score.cell(row=row, column=11, value=f'=IF(E{row}="FILTERED",0,IF(OR(\'2_Inventory_Data\'!I{inv_row}="Transit: Airport",\'2_Inventory_Data\'!I{inv_row}="Transit: Rail/Metro Station"),90,IF(\'2_Inventory_Data\'!I{inv_row}="Retail: Shopping Center",75,65)))')
        
        # Quality Fit (based on daily impressions - higher = better quality placement)
        ws_score.cell(row=row, column=12, value=f'=IF(E{row}="FILTERED",0,MIN(100,\'2_Inventory_Data\'!L{inv_row}/5000))')
        
        # Time Fit (based on operating hours)
        ws_score.cell(row=row, column=13, value=f'=IF(E{row}="FILTERED",0,MIN(100,\'2_Inventory_Data\'!N{inv_row}*4.5))')
        
        # Final Weighted Score (8 factors)
        ws_score.cell(row=row, column=14, value=f'=IF(E{row}="FILTERED",0,ROUND(G{row}*0.15+H{row}*0.10+I{row}*0.18+J{row}*0.22+K{row}*0.10+L{row}*0.10+M{row}*0.10+F{row}*0.05,1))')
        
        # Overall Rank (only for matched items)
        ws_score.cell(row=row, column=15, value=f'=IF(N{row}=0,"",RANK(N{row},$N$5:$N$44,0))')
        
        # Digital Rank (rank only among Digital inventories)
        ws_score.cell(row=row, column=16, value=f'=IF(OR(E{row}="FILTERED",\'2_Inventory_Data\'!F{inv_row}<>"Digital"),"",COUNTIFS($E$5:$E$44,"MATCH",\'2_Inventory_Data\'!$F$2:$F$41,"Digital",$N$5:$N$44,">"&N{row})+1)')
        
        # Classic Rank (rank only among Classic inventories)
        ws_score.cell(row=row, column=17, value=f'=IF(OR(E{row}="FILTERED",\'2_Inventory_Data\'!F{inv_row}<>"Classic"),"",COUNTIFS($E$5:$E$44,"MATCH",\'2_Inventory_Data\'!$F$2:$F$41,"Classic",$N$5:$N$44,">"&N{row})+1)')
        
        # Audio Rank (rank only among Audio inventories)
        ws_score.cell(row=row, column=18, value=f'=IF(OR(E{row}="FILTERED",\'2_Inventory_Data\'!F{inv_row}<>"Audio"),"",COUNTIFS($E$5:$E$44,"MATCH",\'2_Inventory_Data\'!$F$2:$F$41,"Audio",$N$5:$N$44,">"&N{row})+1)')
    
    # Add headers for new rank columns
    ws_score.cell(row=4, column=16, value="Digital Rank")
    ws_score.cell(row=4, column=16).font = header_font
    ws_score.cell(row=4, column=16).fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    
    ws_score.cell(row=4, column=17, value="Classic Rank")
    ws_score.cell(row=4, column=17).font = header_font
    ws_score.cell(row=4, column=17).fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
    
    ws_score.cell(row=4, column=18, value="Audio Rank")
    ws_score.cell(row=4, column=18).font = header_font
    ws_score.cell(row=4, column=18).fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    
    # Add color scale to Final Score
    color_scale = ColorScaleRule(
        start_type='min', start_color='F8696B',
        mid_type='percentile', mid_value=50, mid_color='FFEB84',
        end_type='max', end_color='63BE7B'
    )
    ws_score.conditional_formatting.add('N5:N44', color_scale)
    
    # Column widths
    ws_score.column_dimensions['A'].width = 8
    ws_score.column_dimensions['B'].width = 28
    ws_score.column_dimensions['C'].width = 14
    for col in range(4, 16):
        ws_score.column_dimensions[get_column_letter(col)].width = 12
    
    ws_score.freeze_panes = 'A5'
    
    # ========== SHEET 4: SELECTION ==========
    ws_select = wb.create_sheet("4_Selection")
    
    ws_select['A1'] = "BUDGET ALLOCATION & INVENTORY SELECTION"
    ws_select['A1'].font = Font(bold=True, size=14, color="1F4E79")
    ws_select.merge_cells('A1:I1')
    
    # Budget Allocation Section
    ws_select['A3'] = "BUDGET ALLOCATION BY CLASSIFICATION"
    ws_select['A3'].font = header_font
    ws_select['A3'].fill = header_fill
    ws_select.merge_cells('A3:E3')
    
    ws_select['A4'] = "Classification"
    ws_select['B4'] = "Allocation %"
    ws_select['C4'] = "Allocated Budget"
    ws_select['D4'] = "Inventories Matched"
    ws_select['E4'] = "Currency"
    for col in range(1, 6):
        ws_select.cell(row=4, column=col).font = Font(bold=True)
        ws_select.cell(row=4, column=col).fill = section_fill
    
    ws_select['A5'] = "Digital"
    ws_select['B5'] = "='1_Inputs'!E18"
    ws_select['B5'].number_format = '0%'
    ws_select['C5'] = "='1_Inputs'!F18"
    ws_select['C5'].number_format = '#,##0'
    ws_select['D5'] = '=COUNTIFS(\'3_Scoring\'!$E$5:$E$44,"MATCH",\'2_Inventory_Data\'!$F$2:$F$41,"Digital")'
    ws_select['E5'] = "='1_Inputs'!C5"
    
    ws_select['A6'] = "Classic"
    ws_select['B6'] = "='1_Inputs'!E19"
    ws_select['B6'].number_format = '0%'
    ws_select['C6'] = "='1_Inputs'!F19"
    ws_select['C6'].number_format = '#,##0'
    ws_select['D6'] = '=COUNTIFS(\'3_Scoring\'!$E$5:$E$44,"MATCH",\'2_Inventory_Data\'!$F$2:$F$41,"Classic")'
    ws_select['E6'] = "='1_Inputs'!C5"
    
    ws_select['A7'] = "Audio"
    ws_select['B7'] = "='1_Inputs'!E20"
    ws_select['B7'].number_format = '0%'
    ws_select['C7'] = "='1_Inputs'!F20"
    ws_select['C7'].number_format = '#,##0'
    ws_select['D7'] = '=COUNTIFS(\'3_Scoring\'!$E$5:$E$44,"MATCH",\'2_Inventory_Data\'!$F$2:$F$41,"Audio")'
    ws_select['E7'] = "='1_Inputs'!C5"
    
    # ========== DIGITAL SELECTION (Priority 1) ==========
    ws_select['A10'] = "DIGITAL INVENTORIES (Priority 1 - 60% Budget)"
    ws_select['A10'].font = header_font
    ws_select['A10'].fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    ws_select.merge_cells('A10:I10')
    
    sel_headers = ["Rank", "ID", "Name", "City", "Score", "Daily Cost", "Cumulative", "Budget Cap", "Selected"]
    for col, header in enumerate(sel_headers, 1):
        cell = ws_select.cell(row=11, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = section_fill
    
    # Digital inventories (rows 12-17, up to 6)
    # Uses Digital Rank column (P) from Scoring sheet for simple lookup
    for rank in range(1, 7):
        row = 11 + rank
        ws_select.cell(row=row, column=1, value=rank)
        # Find inventory with this Digital Rank using INDEX/MATCH on column P
        ws_select.cell(row=row, column=2, value=f'=IFERROR(INDEX(\'3_Scoring\'!$A$5:$A$44,MATCH({rank},\'3_Scoring\'!$P$5:$P$44,0)),"")')
        ws_select.cell(row=row, column=3, value=f'=IF(B{row}="","",INDEX(\'3_Scoring\'!$B$5:$B$44,MATCH(B{row},\'3_Scoring\'!$A$5:$A$44,0)))')
        ws_select.cell(row=row, column=4, value=f'=IF(B{row}="","",INDEX(\'3_Scoring\'!$C$5:$C$44,MATCH(B{row},\'3_Scoring\'!$A$5:$A$44,0)))')
        ws_select.cell(row=row, column=5, value=f'=IF(B{row}="","",INDEX(\'3_Scoring\'!$N$5:$N$44,MATCH(B{row},\'3_Scoring\'!$A$5:$A$44,0)))')
        ws_select.cell(row=row, column=6, value=f'=IFERROR(IF(B{row}="","",INDEX(\'2_Inventory_Data\'!$L$2:$L$41,MATCH(B{row},\'2_Inventory_Data\'!$A$2:$A$41,0))/1000*INDEX(\'2_Inventory_Data\'!$J$2:$J$41,MATCH(B{row},\'2_Inventory_Data\'!$A$2:$A$41,0))),"")')
        ws_select[f'F{row}'].number_format = '#,##0'
        ws_select.cell(row=row, column=7, value=f'=IF(B{row}="","",SUM($F$12:F{row})*\'1_Inputs\'!$E$12)')
        ws_select[f'G{row}'].number_format = '#,##0'
        ws_select.cell(row=row, column=8, value=f'=$C$5')  # Digital budget cap
        ws_select[f'H{row}'].number_format = '#,##0'
        ws_select.cell(row=row, column=9, value=f'=IF(B{row}="","",IF(G{row}<=H{row}*1.05,"YES","NO"))')
    
    # ========== CLASSIC SELECTION (Priority 2) ==========
    ws_select['A19'] = "CLASSIC INVENTORIES (Priority 2 - 35% Budget)"
    ws_select['A19'].font = header_font
    ws_select['A19'].fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
    ws_select.merge_cells('A19:I19')
    
    for col, header in enumerate(sel_headers, 1):
        cell = ws_select.cell(row=20, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = section_fill
    
    # Classic inventories (rows 21-25, up to 5)
    # Uses Classic Rank column (Q) from Scoring sheet for simple lookup
    for rank in range(1, 6):
        row = 20 + rank
        ws_select.cell(row=row, column=1, value=rank)
        # Find inventory with this Classic Rank using INDEX/MATCH on column Q
        ws_select.cell(row=row, column=2, value=f'=IFERROR(INDEX(\'3_Scoring\'!$A$5:$A$44,MATCH({rank},\'3_Scoring\'!$Q$5:$Q$44,0)),"")')
        ws_select.cell(row=row, column=3, value=f'=IF(B{row}="","",INDEX(\'3_Scoring\'!$B$5:$B$44,MATCH(B{row},\'3_Scoring\'!$A$5:$A$44,0)))')
        ws_select.cell(row=row, column=4, value=f'=IF(B{row}="","",INDEX(\'3_Scoring\'!$C$5:$C$44,MATCH(B{row},\'3_Scoring\'!$A$5:$A$44,0)))')
        ws_select.cell(row=row, column=5, value=f'=IF(B{row}="","",INDEX(\'3_Scoring\'!$N$5:$N$44,MATCH(B{row},\'3_Scoring\'!$A$5:$A$44,0)))')
        ws_select.cell(row=row, column=6, value=f'=IFERROR(IF(B{row}="","",INDEX(\'2_Inventory_Data\'!$L$2:$L$41,MATCH(B{row},\'2_Inventory_Data\'!$A$2:$A$41,0))/1000*INDEX(\'2_Inventory_Data\'!$J$2:$J$41,MATCH(B{row},\'2_Inventory_Data\'!$A$2:$A$41,0))),"")')
        ws_select[f'F{row}'].number_format = '#,##0'
        ws_select.cell(row=row, column=7, value=f'=IF(B{row}="","",SUM($F$21:F{row})*\'1_Inputs\'!$E$12)')
        ws_select[f'G{row}'].number_format = '#,##0'
        ws_select.cell(row=row, column=8, value=f'=$C$6')  # Classic budget cap
        ws_select[f'H{row}'].number_format = '#,##0'
        ws_select.cell(row=row, column=9, value=f'=IF(B{row}="","",IF(G{row}<=H{row}*1.05,"YES","NO"))')
    
    # ========== AUDIO SELECTION (Priority 3 - Lowest) ==========
    ws_select['A27'] = "AUDIO INVENTORIES (Priority 3 - 5% Budget, Lowest Priority)"
    ws_select['A27'].font = header_font
    ws_select['A27'].fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    ws_select.merge_cells('A27:I27')
    
    for col, header in enumerate(sel_headers, 1):
        cell = ws_select.cell(row=28, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = section_fill
    
    # Audio inventories (rows 29-31, up to 3)
    # Uses Audio Rank column (R) from Scoring sheet for simple lookup
    for rank in range(1, 4):
        row = 28 + rank
        ws_select.cell(row=row, column=1, value=rank)
        # Find inventory with this Audio Rank using INDEX/MATCH on column R
        ws_select.cell(row=row, column=2, value=f'=IFERROR(INDEX(\'3_Scoring\'!$A$5:$A$44,MATCH({rank},\'3_Scoring\'!$R$5:$R$44,0)),"")')
        ws_select.cell(row=row, column=3, value=f'=IF(B{row}="","",INDEX(\'3_Scoring\'!$B$5:$B$44,MATCH(B{row},\'3_Scoring\'!$A$5:$A$44,0)))')
        ws_select.cell(row=row, column=4, value=f'=IF(B{row}="","",INDEX(\'3_Scoring\'!$C$5:$C$44,MATCH(B{row},\'3_Scoring\'!$A$5:$A$44,0)))')
        ws_select.cell(row=row, column=5, value=f'=IF(B{row}="","",INDEX(\'3_Scoring\'!$N$5:$N$44,MATCH(B{row},\'3_Scoring\'!$A$5:$A$44,0)))')
        ws_select.cell(row=row, column=6, value=f'=IFERROR(IF(B{row}="","",INDEX(\'2_Inventory_Data\'!$L$2:$L$41,MATCH(B{row},\'2_Inventory_Data\'!$A$2:$A$41,0))/1000*INDEX(\'2_Inventory_Data\'!$J$2:$J$41,MATCH(B{row},\'2_Inventory_Data\'!$A$2:$A$41,0))),"")')
        ws_select[f'F{row}'].number_format = '#,##0'
        ws_select.cell(row=row, column=7, value=f'=IF(B{row}="","",SUM($F$29:F{row})*\'1_Inputs\'!$E$12)')
        ws_select[f'G{row}'].number_format = '#,##0'
        ws_select.cell(row=row, column=8, value=f'=$C$7')  # Audio budget cap
        ws_select[f'H{row}'].number_format = '#,##0'
        # Audio only selected if: within audio budget AND audio allocation > 0
        ws_select.cell(row=row, column=9, value=f'=IF(B{row}="","",IF(AND(G{row}<=H{row}*1.05,$B$7>0),"YES","NO"))')
    
    # Note about audio
    ws_select['A32'] = "Note: Audio is lowest priority. Only selected when allocation > 0% and within budget cap."
    ws_select['A32'].font = Font(italic=True, size=9, color="7030A0")
    ws_select.merge_cells('A32:I32')
    
    # Add conditional formatting for selected items
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ws_select.conditional_formatting.add('I12:I17', FormulaRule(formula=['$I12="YES"'], fill=green_fill))
    ws_select.conditional_formatting.add('I21:I25', FormulaRule(formula=['$I21="YES"'], fill=green_fill))
    ws_select.conditional_formatting.add('I29:I31', FormulaRule(formula=['$I29="YES"'], fill=green_fill))
    
    # Column widths
    ws_select.column_dimensions['A'].width = 10
    ws_select.column_dimensions['B'].width = 10
    ws_select.column_dimensions['C'].width = 28
    ws_select.column_dimensions['D'].width = 14
    ws_select.column_dimensions['E'].width = 14
    ws_select.column_dimensions['F'].width = 10
    ws_select.column_dimensions['G'].width = 14
    ws_select.column_dimensions['H'].width = 14
    ws_select.column_dimensions['I'].width = 10
    
    # ========== SHEET 5: SCHEDULES ==========
    ws_sched = wb.create_sheet("5_Schedules")
    
    ws_sched['A1'] = "SCHEDULE CREATION - DETAILED VIEW"
    ws_sched['A1'].font = Font(bold=True, size=14, color="1F4E79")
    ws_sched.merge_cells('A1:O1')
    
    ws_sched['A2'] = "Each selected inventory gets a schedule with start/end dates and optimal time slots based on operating hours."
    ws_sched['A2'].font = Font(italic=True, size=9)
    ws_sched.merge_cells('A2:O2')
    
    # Schedule headers - expanded with dates and times
    sched_headers = [
        "Inventory ID", "Name", "City", "Classification", 
        "Start Date", "Start Time", "End Date", "End Time",
        "Op. Hours", "Hours/Day", "Daily Impr.", "Daily Cost",
        "Campaign Impr.", "Campaign Cost", "Contrib. to Goal"
    ]
    for col, header in enumerate(sched_headers, 1):
        cell = ws_sched.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Define selection row mappings (Digital: 12-17, Classic: 21-25, Audio: 29-31)
    selection_rows = [12, 13, 14, 15, 16, 17, 21, 22, 23, 24, 25, 29, 30, 31]
    
    # Schedule formulas for selected inventories
    for idx, sel_row in enumerate(selection_rows[:14]):  # Up to 14 inventories
        row = 5 + idx
        
        # Inventory ID (only show if selected)
        ws_sched.cell(row=row, column=1, value=f"=IF('4_Selection'!I{sel_row}=\"YES\",'4_Selection'!B{sel_row},\"\")")
        
        # Name
        ws_sched.cell(row=row, column=2, value=f"=IF(A{row}=\"\",\"\",'4_Selection'!C{sel_row})")
        
        # City
        ws_sched.cell(row=row, column=3, value=f"=IF(A{row}=\"\",\"\",'4_Selection'!D{sel_row})")
        
        # Classification
        ws_sched.cell(row=row, column=4, value=f'=IF(A{row}="","",INDEX(\'2_Inventory_Data\'!$F$2:$F$41,MATCH(A{row},\'2_Inventory_Data\'!$A$2:$A$41,0)))')
        
        # Start Date (from campaign inputs)
        ws_sched.cell(row=row, column=5, value=f'=IF(A{row}="","",\'1_Inputs\'!$B$7)')
        
        # Start Time (based on operating hours - assume starts at 6:00 or 8:00)
        ws_sched.cell(row=row, column=6, value=f'=IF(A{row}="","",IF(INDEX(\'2_Inventory_Data\'!$N$2:$N$41,MATCH(A{row},\'2_Inventory_Data\'!$A$2:$A$41,0))>=20,"06:00",IF(INDEX(\'2_Inventory_Data\'!$N$2:$N$41,MATCH(A{row},\'2_Inventory_Data\'!$A$2:$A$41,0))>=16,"07:00","08:00")))')
        
        # End Date (from campaign inputs)
        ws_sched.cell(row=row, column=7, value=f'=IF(A{row}="","",\'1_Inputs\'!$B$8)')
        
        # End Time (start time + operating hours, capped at 24:00)
        ws_sched.cell(row=row, column=8, value=f'=IF(A{row}="","",IF(INDEX(\'2_Inventory_Data\'!$N$2:$N$41,MATCH(A{row},\'2_Inventory_Data\'!$A$2:$A$41,0))=24,"23:59",TEXT(MIN(23,VALUE(LEFT(F{row},2))+MIN(12,INDEX(\'2_Inventory_Data\'!$N$2:$N$41,MATCH(A{row},\'2_Inventory_Data\'!$A$2:$A$41,0)))),"00")&":00"))')
        
        # Operating Hours (from inventory)
        ws_sched.cell(row=row, column=9, value=f'=IF(A{row}="","",INDEX(\'2_Inventory_Data\'!$N$2:$N$41,MATCH(A{row},\'2_Inventory_Data\'!$A$2:$A$41,0)))')
        
        # Hours/Day scheduled (capped at 12 for budget efficiency)
        ws_sched.cell(row=row, column=10, value=f'=IF(A{row}="","",MIN(I{row},12))')
        
        # Daily Impressions
        ws_sched.cell(row=row, column=11, value=f'=IF(A{row}="","",ROUND(INDEX(\'2_Inventory_Data\'!$L$2:$L$41,MATCH(A{row},\'2_Inventory_Data\'!$A$2:$A$41,0))*(J{row}/I{row}),0))')
        ws_sched[f'K{row}'].number_format = '#,##0'
        
        # Daily Cost
        ws_sched.cell(row=row, column=12, value=f"=IF(A{row}=\"\",\"\",'4_Selection'!G{sel_row})")
        ws_sched[f'L{row}'].number_format = '#,##0'
        
        # Campaign Impressions (daily * days)
        ws_sched.cell(row=row, column=13, value=f'=IF(A{row}="","",K{row}*\'1_Inputs\'!$E$12)')
        ws_sched[f'M{row}'].number_format = '#,##0'
        
        # Campaign Cost (daily * days)
        ws_sched.cell(row=row, column=14, value=f'=IF(A{row}="","",L{row}*\'1_Inputs\'!$E$12)')
        ws_sched[f'N{row}'].number_format = '#,##0'
        
        # Contribution to Goal (%)
        ws_sched.cell(row=row, column=15, value=f'=IF(A{row}="","",IFERROR(M{row}/\'1_Inputs\'!$B$11,0))')
        ws_sched[f'O{row}'].number_format = '0.0%'
    
    # ========== CAMPAIGN SUMMARY ==========
    ws_sched['A17'] = "CAMPAIGN SUMMARY"
    ws_sched['A17'].font = header_font
    ws_sched['A17'].fill = header_fill
    ws_sched.merge_cells('A17:D17')
    
    # Summary metrics
    ws_sched['A18'] = "Metric"
    ws_sched['B18'] = "Value"
    ws_sched['C18'] = "Target"
    ws_sched['D18'] = "Status"
    for col in ['A', 'B', 'C', 'D']:
        ws_sched[f'{col}18'].font = Font(bold=True)
        ws_sched[f'{col}18'].fill = section_fill
    
    # Total Impressions
    ws_sched['A19'] = "Total Impressions"
    ws_sched['B19'] = "=SUM(M5:M14)"
    ws_sched['B19'].number_format = '#,##0'
    ws_sched['C19'] = "='1_Inputs'!B11"
    ws_sched['C19'].number_format = '#,##0'
    ws_sched['D19'] = '=IF(B19>=C19*1.1,"EXCEEDED",IF(B19>=C19*0.95,"MET",IF(B19>=C19*0.8,"CLOSE","BELOW")))'
    
    # Total Cost
    ws_sched['A20'] = "Total Cost"
    ws_sched['B20'] = "=SUM(N5:N14)"
    ws_sched['B20'].number_format = '#,##0'
    ws_sched['C20'] = "='1_Inputs'!B9"
    ws_sched['C20'].number_format = '#,##0'
    ws_sched['D20'] = '=IF(B20<=C20*0.95,"UNDER BUDGET",IF(B20<=C20*1.05,"ON BUDGET","OVER BUDGET"))'
    
    # Goal Achievement %
    ws_sched['A21'] = "Goal Achievement"
    ws_sched['B21'] = "=IFERROR(B19/C19,0)"
    ws_sched['B21'].number_format = '0.0%'
    ws_sched['C21'] = "100%"
    ws_sched['D21'] = '=IF(B21>=1.1,"EXCEEDED TARGET",IF(B21>=0.95,"TARGET MET",IF(B21>=0.8,"NEARLY MET","BELOW TARGET")))'
    
    # Budget Utilization %
    ws_sched['A22'] = "Budget Utilization"
    ws_sched['B22'] = "=IFERROR(B20/C20,0)"
    ws_sched['B22'].number_format = '0.0%'
    ws_sched['C22'] = "100%"
    ws_sched['D22'] = '=IF(B22>=1.05,"OVER BUDGET",IF(B22>=0.95,"OPTIMAL","UNDER-UTILIZED"))'
    
    # Inventories Selected
    ws_sched['A23'] = "Inventories Selected"
    ws_sched['B23'] = '=COUNTIF(\'4_Selection\'!I12:I26,"YES")'
    ws_sched['C23'] = ""
    ws_sched['D23'] = ""
    
    # Efficiency Metrics
    ws_sched['A25'] = "EFFICIENCY METRICS"
    ws_sched['A25'].font = header_font
    ws_sched['A25'].fill = header_fill
    ws_sched.merge_cells('A25:D25')
    
    ws_sched['A26'] = "Effective CPM"
    ws_sched['B26'] = "=IFERROR(B20/(B19/1000),0)"
    ws_sched['B26'].number_format = '#,##0.00'
    ws_sched['C26'] = "='1_Inputs'!C5"
    ws_sched['D26'] = '=IF(B26<20,"EXCELLENT",IF(B26<50,"GOOD",IF(B26<100,"AVERAGE","EXPENSIVE")))'
    
    ws_sched['A27'] = "Cost per Day"
    ws_sched['B27'] = "=IFERROR(B20/'1_Inputs'!E12,0)"
    ws_sched['B27'].number_format = '#,##0'
    ws_sched['C27'] = "='1_Inputs'!C5"
    ws_sched['D27'] = ""
    
    ws_sched['A28'] = "Impressions per Day"
    ws_sched['B28'] = "=IFERROR(B19/'1_Inputs'!E12,0)"
    ws_sched['B28'].number_format = '#,##0'
    ws_sched['C28'] = ""
    ws_sched['D28'] = ""
    
    # Conditional formatting for status
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    ws_sched.conditional_formatting.add('D19:D22', FormulaRule(formula=['OR($D19="MET",$D19="ON BUDGET",$D19="TARGET MET",$D19="OPTIMAL")'], fill=green_fill))
    ws_sched.conditional_formatting.add('D19:D22', FormulaRule(formula=['OR($D19="EXCEEDED",$D19="EXCEEDED TARGET")'], fill=green_fill))
    ws_sched.conditional_formatting.add('D19:D22', FormulaRule(formula=['OR($D19="CLOSE",$D19="NEARLY MET",$D19="UNDER BUDGET",$D19="UNDER-UTILIZED")'], fill=yellow_fill))
    ws_sched.conditional_formatting.add('D19:D22', FormulaRule(formula=['OR($D19="BELOW",$D19="BELOW TARGET",$D19="OVER BUDGET")'], fill=red_fill))
    
    # Column widths
    ws_sched.column_dimensions['A'].width = 12
    ws_sched.column_dimensions['B'].width = 26
    ws_sched.column_dimensions['C'].width = 14
    ws_sched.column_dimensions['D'].width = 12
    ws_sched.column_dimensions['E'].width = 12
    ws_sched.column_dimensions['F'].width = 11
    ws_sched.column_dimensions['G'].width = 12
    ws_sched.column_dimensions['H'].width = 11
    ws_sched.column_dimensions['I'].width = 10
    ws_sched.column_dimensions['J'].width = 10
    ws_sched.column_dimensions['K'].width = 12
    ws_sched.column_dimensions['L'].width = 12
    ws_sched.column_dimensions['M'].width = 14
    ws_sched.column_dimensions['N'].width = 14
    ws_sched.column_dimensions['O'].width = 14
    
    # ========== SHEET 6: VISUALIZATION ==========
    ws_viz = wb.create_sheet("6_Visualization")
    
    ws_viz['A1'] = "CAMPAIGN PLAN VISUALIZATION"
    ws_viz['A1'].font = Font(bold=True, size=14, color="1F4E79")
    ws_viz.merge_cells('A1:H1')
    
    # Summary metrics
    ws_viz['A3'] = "CAMPAIGN PARAMETERS"
    ws_viz['A3'].font = header_font
    ws_viz['A3'].fill = header_fill
    ws_viz.merge_cells('A3:B3')
    
    params = [
        ("Country", "='1_Inputs'!B5"),
        ("City Filter", "='1_Inputs'!B6"),
        ("Venue Filter", "='1_Inputs'!B14"),
        ("Campaign Duration", "='1_Inputs'!E12&\" days\""),
        ("Budget", "='1_Inputs'!C5&\" \"&TEXT('1_Inputs'!B9,\"#,##0\")"),
        ("Goal Type", "='1_Inputs'!B10"),
        ("Goal Value", "=TEXT('1_Inputs'!B11,\"#,##0\")"),
        ("Pricing Model", "='1_Inputs'!E13"),
    ]
    
    for idx, (label, formula) in enumerate(params, 4):
        ws_viz.cell(row=idx, column=1, value=label)
        ws_viz.cell(row=idx, column=1).font = label_font
        ws_viz.cell(row=idx, column=2, value=formula)
    
    # Results
    ws_viz['A13'] = "RESULTS"
    ws_viz['A13'].font = header_font
    ws_viz['A13'].fill = header_fill
    ws_viz.merge_cells('A13:B13')
    
    results = [
        ("Inventories Selected", "=COUNTIF('4_Selection'!I12:I26,\"YES\")"),
        ("Total Impressions", "='5_Schedules'!B16"),
        ("Total Cost", "='1_Inputs'!C5&\" \"&TEXT('5_Schedules'!B17,\"#,##0\")"),
        ("Budget Utilization", "=TEXT('5_Schedules'!B18,\"0.0%\")"),
        ("Goal Achievement", "=TEXT('5_Schedules'!B19,\"0.0%\")"),
    ]
    
    for idx, (label, formula) in enumerate(results, 14):
        ws_viz.cell(row=idx, column=1, value=label)
        ws_viz.cell(row=idx, column=1).font = label_font
        ws_viz.cell(row=idx, column=2, value=formula)
    
    # Budget allocation by classification (for chart)
    ws_viz['D3'] = "ALLOCATION BY CLASSIFICATION"
    ws_viz['D3'].font = header_font
    ws_viz['D3'].fill = header_fill
    ws_viz.merge_cells('D3:E3')
    
    ws_viz['D4'] = "Classification"
    ws_viz['E4'] = "Budget"
    ws_viz['D5'] = "Digital"
    ws_viz['E5'] = "='1_Inputs'!F18"
    ws_viz['D6'] = "Classic"
    ws_viz['E6'] = "='1_Inputs'!F19"
    ws_viz['D7'] = "Audio"
    ws_viz['E7'] = "='1_Inputs'!F20"
    
    # Create pie chart
    pie = PieChart()
    pie.title = "Budget Allocation by Classification"
    data = Reference(ws_viz, min_col=5, min_row=4, max_row=7)
    cats = Reference(ws_viz, min_col=4, min_row=5, max_row=7)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(cats)
    pie.width = 14
    pie.height = 10
    ws_viz.add_chart(pie, "D10")
    
    # Column widths
    ws_viz.column_dimensions['A'].width = 22
    ws_viz.column_dimensions['B'].width = 22
    ws_viz.column_dimensions['D'].width = 18
    ws_viz.column_dimensions['E'].width = 15
    
    # ========== SHEET 7: FORMULA REFERENCE ==========
    ws_ref = wb.create_sheet("7_Formula_Reference")
    
    ws_ref['A1'] = "FORMULA REFERENCE - Understanding the Calculations"
    ws_ref['A1'].font = Font(bold=True, size=14, color="1F4E79")
    ws_ref.merge_cells('A1:D1')
    
    ws_ref['A3'] = "This sheet explains all formulas used in the simulator. Click any cell in other sheets to see the formula."
    ws_ref['A3'].font = Font(italic=True)
    ws_ref.merge_cells('A3:D3')
    
    formulas = [
        ("Section", "Formula Name", "Formula Logic", "Explanation"),
        ("Inputs", "Currency", "=IF(Country=\"Malaysia\",\"MYR\",\"JPY\")", "Auto-selects currency based on country"),
        ("Inputs", "Pricing Model", "=IF(Goal IN (Impr,Reach,CO2),\"CPM\",\"CPS\")", "CPM for impressions/reach, CPS for plays/SOV"),
        ("Inputs", "Digital Allocation", "=IF(Goal=\"Ad Plays\",95%,...)", "Goal-driven allocation (see PRD Section 12.3)"),
        ("Scoring", "Filter Match", "=IF(Country+City+Venue+Class match,\"MATCH\",\"FILTERED\")", "Only matched inventories get scored"),
        ("Scoring", "Geo Fit", "=IF(Tier=1,100,IF(Tier=2,85,70))", "Higher tier cities get higher geo scores"),
        ("Scoring", "Budget Fit", "=MAX(0,100-CPM*0.3)", "Lower CPM = more budget efficient = higher score"),
        ("Scoring", "Audience Fit", "=(Age_Match+Interest_Match+Behavior_Match)/3", "Weighted average of audience segment matches"),
        ("Scoring", "Final Score", "=Weighted sum of 8 factors", "Weights: Geo 15%, Avail 10%, Budget 18%, Audience 22%, Brand 10%, Quality 10%, Time 10%, Country 5%"),
        ("Selection", "Daily Cost (CPM)", "=Daily_Impressions/1000 * CPM", "Cost per thousand impressions * volume"),
        ("Selection", "Daily Cost (CPS)", "=CPS * Slots_Per_Loop * Operating_Hours", "Cost per slot * slots per hour * hours"),
        ("Selection", "Selected", "=IF(Cumulative_Cost <= Budget*1.05,\"YES\",\"NO\")", "5% tolerance for budget overage"),
        ("Schedules", "Campaign Cost", "=Daily_Cost * Campaign_Days", "Total cost = daily rate * duration"),
        ("Schedules", "Goal Achievement", "=Total_Impressions / Goal_Value", "How close to target goal"),
        ("City Tier", "Tier 1 Cities", "MY: Kuala Lumpur, KLIA | JP: Tokyo, Osaka", "Capital regions, 10M+ population"),
        ("City Tier", "Tier 2 Cities", "MY: Penang, JB, Ipoh | JP: Nagoya, Yokohama, Kyoto", "Regional capitals, 1-5M population"),
        ("City Tier", "Tier 3 Cities", "MY: Melaka, etc. | JP: Others", "Smaller cities, <1M population"),
    ]
    
    for row_idx, row_data in enumerate(formulas, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_ref.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 5:
                cell.font = header_font
                cell.fill = header_fill
    
    ws_ref.column_dimensions['A'].width = 15
    ws_ref.column_dimensions['B'].width = 22
    ws_ref.column_dimensions['C'].width = 55
    ws_ref.column_dimensions['D'].width = 50
    
    # Save workbook
    output_path = "docs/mw-planner-recommendation-simulator.xlsx"
    wb.save(output_path)
    print(f"Simulator saved to: {output_path}")
    print("\nSheets created:")
    print("1. 1_Inputs - Campaign parameters with ALL dropdowns")
    print("   - Country (Malaysia/Japan)")
    print("   - City filter (All Cities + specific cities)")
    print("   - Venue Type filter (IAB venue types)")
    print("   - Classification filter (Digital/Classic/Audio)")
    print("   - Type filter (OOH/Transit/Retail/Network/Radio)")
    print("   - Audience: Age, Gender, Income, Interest, Behavior")
    print("2. 2_Inventory_Data - Malaysia & Japan inventory (40 items) with City Tier")
    print("3. 3_Scoring - 8-factor scoring with FILTER MATCH logic")
    print("4. 4_Selection - Budget allocation & inventory selection")
    print("5. 5_Schedules - Schedule creation")
    print("6. 6_Visualization - Charts and summary")
    print("7. 7_Formula_Reference - All formulas explained + City Tier logic")

if __name__ == "__main__":
    create_simulator()
